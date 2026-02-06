import csv
import os
import re
import unicodedata
import uuid
from datetime import datetime
from io import BytesIO, StringIO

import openpyxl
import pandas as pd
from flask import Blueprint, flash, redirect, render_template, request, send_file, session, url_for

from core.audit import log_change
from core.config import BASE_DIR
from core.db import get_db
from core.security import login_required, write_required
from core.utils import clean_note, get_appreciation_dynamique

bp = Blueprint("main", __name__)

IMPORT_MAPPING_FIELDS = [
    ("full_name", "Nom complet"),
    ("last_name", "Nom"),
    ("first_name", "Prenom"),
    ("classe", "Classe"),
    ("devoir", "Devoir (/20)"),
    ("activite", "Activite (/20)"),
    ("compo", "Compo (/20)"),
    ("participation", "Participation (/3)"),
    ("comportement", "Comportement (/6)"),
    ("cahier", "Cahier (/5)"),
    ("projet", "Projet (/4)"),
    ("assiduite_outils", "Absences/Outils (/2)"),
    ("remarques", "Remarques"),
    ("phone", "Telephone parent"),
    ("email", "Email parent"),
]


def _parse_float(value):
    try:
        if value is None or value == "":
            return None
        return float(value)
    except Exception:
        return None


def _safe_list_get(values, idx, default=""):
    if idx < len(values):
        return values[idx]
    return default


def _clean_component(value, maximum):
    score = clean_note(value)
    if score > maximum:
        score = float(maximum)
    return round(score, 2)


def _sum_activite_components(participation, comportement, cahier, projet, assiduite_outils):
    return round(
        float(participation)
        + float(comportement)
        + float(cahier)
        + float(projet)
        + float(assiduite_outils),
        2,
    )


def _split_activite_components(total):
    remaining = clean_note(total)
    caps = (3.0, 6.0, 5.0, 4.0, 2.0)
    values = []
    for cap in caps:
        take = min(remaining, cap)
        values.append(round(take, 2))
        remaining = round(max(0.0, remaining - take), 2)
    return tuple(values)


def _normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.lower()
    text = re.sub(r"[\s\-_]+", "", text)
    return text


def _find_header_row(df):
    for i, row in df.head(30).iterrows():
        joined = " ".join([str(v) for v in row.values])
        cells = [_normalize_header(v) for v in row.values]
        if (
            "Nom" in joined
            or "Prenom" in joined
            or "nomcomplet" in "".join(cells)
            or any(c in {"nom", "prenom", "nomcomplet", "nomprenom", "fullname"} for c in cells)
        ):
            return i
    return -1


def _build_default_mapping(columns):
    defaults = {
        "full_name": "",
        "last_name": "",
        "first_name": "",
        "classe": "",
        "devoir": "",
        "activite": "",
        "compo": "",
        "participation": "",
        "comportement": "",
        "cahier": "",
        "projet": "",
        "assiduite_outils": "",
        "remarques": "",
        "phone": "",
        "email": "",
    }

    for col in columns:
        norm = _normalize_header(col)
        if not defaults["full_name"] and norm in ("nomcomplet", "nomprenom", "fullname"):
            defaults["full_name"] = col
        elif not defaults["last_name"] and norm in ("nom", "lastname", "surname"):
            defaults["last_name"] = col
        elif not defaults["first_name"] and norm in ("prenom", "firstname"):
            defaults["first_name"] = col
        elif not defaults["classe"] and norm in ("classe", "niveau", "class"):
            defaults["classe"] = col
        elif not defaults["devoir"] and norm in ("devoir", "dev", "homework", "04", "4"):
            defaults["devoir"] = col
        elif not defaults["activite"] and norm in ("activite", "act", "activity", "01", "1"):
            defaults["activite"] = col
        elif not defaults["compo"] and norm in ("compo", "composition", "exam", "test", "09", "9"):
            defaults["compo"] = col
        elif not defaults["participation"] and norm in ("participation", "moucharaka", "moucharakah"):
            defaults["participation"] = col
        elif not defaults["comportement"] and norm in ("comportement", "souk", "conduite", "behavior"):
            defaults["comportement"] = col
        elif not defaults["cahier"] and norm in ("cahier", "korras", "kras", "copybook"):
            defaults["cahier"] = col
        elif not defaults["projet"] and norm in ("projet", "project"):
            defaults["projet"] = col
        elif not defaults["assiduite_outils"] and norm in ("absencesoutils", "assiduiteoutils", "absoutils", "outils"):
            defaults["assiduite_outils"] = col
        elif not defaults["remarques"] and norm in ("remarques", "appreciation", "rem", "obs", "commentaire"):
            defaults["remarques"] = col
        elif not defaults["phone"] and norm in ("telephone", "tel", "phone", "mobile"):
            defaults["phone"] = col
        elif not defaults["email"] and norm in ("email", "mail", "e-mail"):
            defaults["email"] = col
    return defaults


def _preview_dir():
    path = os.path.join(BASE_DIR, "tmp_imports")
    os.makedirs(path, exist_ok=True)
    return path


def _cleanup_import_previews(max_age_seconds=24 * 3600):
    folder = _preview_dir()
    now = int(datetime.now().timestamp())
    for name in os.listdir(folder):
        full = os.path.join(folder, name)
        try:
            if not os.path.isfile(full):
                continue
            if now - int(os.path.getmtime(full)) > max_age_seconds:
                os.remove(full)
        except Exception:
            continue


def _get_preview_meta(token):
    data = session.get("import_preview") or {}
    if data.get("token") != token:
        return None
    path = data.get("path") or ""
    if not path or not os.path.exists(path):
        return None
    return data


def _clear_preview_meta(meta):
    try:
        path = meta.get("path") or ""
        if path and os.path.exists(path):
            os.remove(path)
    except Exception:
        pass
    session.pop("import_preview", None)


def _row_value(row, column_name):
    if not column_name:
        return None
    if column_name not in row.index:
        return None
    val = row.get(column_name, None)
    return None if pd.isna(val) else val


def _prepare_import_dataframe(raw_df):
    if raw_df is None or raw_df.empty:
        return None, None, False

    header_row = _find_header_row(raw_df)
    header_detected = header_row >= 0
    if header_row < 0:
        header_row = 0

    work = raw_df.copy()
    work.columns = work.iloc[header_row]
    work = work.iloc[header_row + 1 :].dropna(how="all").copy()

    # Ensure unique, non-empty column names for robust mapping.
    normalized = []
    seen = {}
    for col in list(work.columns):
        label = str(col).strip() if col is not None else ""
        if not label:
            label = "col"
        base = label
        idx = 2
        while label in seen:
            label = f"{base}_{idx}"
            idx += 1
        seen[label] = 1
        normalized.append(label)
    work.columns = normalized

    return work, int(header_row), header_detected


def _resolve_mapped_column(columns, selected):
    if not selected:
        return ""
    if selected in columns:
        return selected
    target = _normalize_header(selected)
    for col in columns:
        if _normalize_header(col) == target:
            return col
    return ""


def _parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except Exception:
        return None


def _school_year(now):
    if now.month >= 9:
        return f"{now.year}/{now.year + 1}"
    return f"{now.year - 1}/{now.year}"


def _arabize(value):
    if value is None:
        return ""
    text = str(value)
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display

        reshaped = arabic_reshaper.reshape(text)
        return get_display(reshaped)
    except Exception:
        return text


def _build_filters(user_id, trim, args, moy_expr_override=None):
    niveau = args.get("niveau", "")
    search = (args.get("recherche") or "").strip()
    sort = args.get("sort", "class")
    order = args.get("order", "asc")
    etat = args.get("etat", "all")
    min_moy = _parse_float(args.get("min_moy", ""))
    max_moy = _parse_float(args.get("max_moy", ""))
    if min_moy is not None and max_moy is not None and min_moy > max_moy:
        min_moy, max_moy = max_moy, min_moy

    if order not in ("asc", "desc"):
        order = "asc"

    moy_expr = (
        moy_expr_override
        if moy_expr_override
        else f"((devoir_t{trim} + activite_t{trim})/2.0 + (compo_t{trim}*2.0))/3.0"
    )
    where = "e.user_id = ?"
    params = [user_id]

    if niveau and niveau != "all":
        where += " AND e.niveau = ?"
        params.append(niveau)
    if search:
        where += " AND e.nom_complet LIKE ?"
        params.append(f"%{search}%")

    if min_moy is not None:
        where += f" AND {moy_expr} >= ?"
        params.append(min_moy)
    if max_moy is not None:
        where += f" AND {moy_expr} <= ?"
        params.append(max_moy)

    if etat == "admis":
        where += f" AND {moy_expr} >= 10"
    elif etat == "echec":
        where += f" AND {moy_expr} > 0 AND {moy_expr} < 10"
    elif etat == "non_saisi":
        where += f" AND {moy_expr} <= 0"

    return {
        "niveau": niveau,
        "search": search,
        "sort": sort,
        "order": order,
        "etat": etat,
        "min_moy": min_moy,
        "max_moy": max_moy,
        "moy_expr": moy_expr,
        "where": where,
        "params": params,
    }


def _build_history_filters(user_id, args):
    action = (args.get("action") or "").strip()
    q = (args.get("q") or "").strip()
    subject_val = (args.get("subject") or "").strip()
    date_from_raw = (args.get("from") or "").strip()
    date_to_raw = (args.get("to") or "").strip()

    subject_id = None
    try:
        if subject_val:
            subject_id = int(subject_val)
    except Exception:
        subject_id = None

    date_from = _parse_date(date_from_raw)
    date_to = _parse_date(date_to_raw)

    where = "l.user_id = ?"
    params = [user_id]

    if action:
        where += " AND l.action = ?"
        params.append(action)
    if subject_id:
        where += " AND l.subject_id = ?"
        params.append(subject_id)
    if q:
        where += " AND (l.details LIKE ? OR e.nom_complet LIKE ?)"
        params.extend([f"%{q}%", f"%{q}%"])
    if date_from:
        where += " AND l.created_at >= ?"
        params.append(int(date_from.timestamp()))
    if date_to:
        end = date_to.replace(hour=23, minute=59, second=59)
        where += " AND l.created_at <= ?"
        params.append(int(end.timestamp()))

    return {
        "where": where,
        "params": params,
        "action": action,
        "q": q,
        "subject_id": subject_id,
        "date_from": date_from_raw,
        "date_to": date_to_raw,
    }


def _get_subjects(db, user_id):
    default_subject = ""
    lock_subject = 0
    try:
        user = db.execute(
            "SELECT default_subject, lock_subject FROM users WHERE id = ?",
            (user_id,),
        ).fetchone()
        default_subject = (user["default_subject"] or "").strip() if user else ""
        lock_subject = int(user["lock_subject"] or 0) if user else 0
    except Exception:
        pass

    rows = db.execute(
        "SELECT id, name FROM subjects WHERE user_id = ? ORDER BY name",
        (user_id,),
    ).fetchall()

    if lock_subject and not default_subject:
        if rows:
            default_subject = (rows[0]["name"] or "").strip()
        else:
            default_subject = "Sciences"
        try:
            db.execute(
                "UPDATE users SET default_subject = ? WHERE id = ?",
                (default_subject, user_id),
            )
            db.commit()
        except Exception:
            pass

    if not rows:
        db.execute(
            "INSERT INTO subjects (user_id, name) VALUES (?, ?)",
            (user_id, default_subject or "Sciences"),
        )
        db.commit()
        rows = db.execute(
            "SELECT id, name FROM subjects WHERE user_id = ? ORDER BY name",
            (user_id,),
        ).fetchall()

    if lock_subject and default_subject:
        match = None
        for r in rows:
            if (r["name"] or "").strip().lower() == default_subject.lower():
                match = r
                break
        if not match:
            db.execute(
                "INSERT OR IGNORE INTO subjects (user_id, name) VALUES (?, ?)",
                (user_id, default_subject),
            )
            db.commit()
            rows = db.execute(
                "SELECT id, name FROM subjects WHERE user_id = ? ORDER BY name",
                (user_id,),
            ).fetchall()
            for r in rows:
                if (r["name"] or "").strip().lower() == default_subject.lower():
                    match = r
                    break
        if match:
            rows = [match]
    return rows


def _select_subject_id(subjects, value):
    subject_map = {int(s["id"]): s for s in subjects}
    try:
        sid = int(value)
    except Exception:
        sid = None
    if sid not in subject_map:
        sid = int(subjects[0]["id"])
    return sid


def _note_expr(trim):
    devoir = f"COALESCE(n.devoir, e.devoir_t{trim})"
    activite = f"COALESCE(n.activite, e.activite_t{trim})"
    compo = f"COALESCE(n.compo, e.compo_t{trim})"
    remarques = f"COALESCE(n.remarques, e.remarques_t{trim})"
    moy_expr = f"(({devoir} + {activite})/2.0 + ({compo}*2.0))/3.0"
    return devoir, activite, compo, remarques, moy_expr


def _parse_trim(raw, default="1"):
    trim = str(raw or default).strip()
    if trim not in ("1", "2", "3"):
        return default
    return trim


@bp.route("/lang/<lang>")
def set_lang(lang: str):
    selected = (lang or "").strip().lower()
    if selected not in ("fr", "ar"):
        selected = "fr"
    session["lang"] = selected
    if session.get("user_id"):
        return redirect(request.referrer or url_for("main.index"))
    return redirect(request.referrer or url_for("auth.login"))


def _build_bulletin_multisubject(db, user_id: int, eleve_id: int, trim: str):
    eleve = db.execute(
        "SELECT * FROM eleves WHERE id = ? AND user_id = ?",
        (eleve_id, user_id),
    ).fetchone()
    if not eleve:
        return None

    subjects = db.execute(
        "SELECT id, name FROM subjects WHERE user_id = ? ORDER BY name COLLATE NOCASE",
        (user_id,),
    ).fetchall()
    if not subjects:
        db.execute(
            "INSERT INTO subjects (user_id, name) VALUES (?, ?)",
            (user_id, "Sciences"),
        )
        db.commit()

    devoir_expr, activite_expr, compo_expr, remarques_expr, _ = _note_expr(trim)

    rows = db.execute(
        f"""
        SELECT
            s.id AS subject_id,
            s.name AS subject_name,
            {activite_expr} AS activite,
            {devoir_expr} AS devoir,
            {compo_expr} AS compo,
            {remarques_expr} AS remarques,
            ROUND((({devoir_expr} + {activite_expr})/2.0 + ({compo_expr} * 2.0))/3.0, 2) AS moyenne
        FROM subjects s
        LEFT JOIN eleves e
            ON e.id = ?
           AND e.user_id = s.user_id
        LEFT JOIN notes n
            ON n.user_id = s.user_id
           AND n.eleve_id = e.id
           AND n.subject_id = s.id
           AND n.trimestre = ?
        WHERE s.user_id = ?
        ORDER BY s.name COLLATE NOCASE
        """,
        (eleve_id, int(trim), user_id),
    ).fetchall()

    subject_lines = []
    for r in rows:
        subject_lines.append(
            {
                "subject_id": int(r["subject_id"]),
                "subject_name": r["subject_name"],
                "activite": float(r["activite"] or 0),
                "devoir": float(r["devoir"] or 0),
                "compo": float(r["compo"] or 0),
                "remarques": r["remarques"] or "",
                "moyenne": float(r["moyenne"] or 0),
            }
        )

    moyenne_generale = 0.0
    if subject_lines:
        moyenne_generale = round(
            sum(line["moyenne"] for line in subject_lines) / len(subject_lines),
            2,
        )

    class_rows = db.execute(
        f"""
        SELECT
            e.id,
            AVG(((COALESCE(n.devoir, e.devoir_t{trim}) + COALESCE(n.activite, e.activite_t{trim}))/2.0 + (COALESCE(n.compo, e.compo_t{trim}) * 2.0))/3.0) AS moyenne_generale
        FROM eleves e
        JOIN subjects s ON s.user_id = e.user_id
        LEFT JOIN notes n
            ON n.user_id = e.user_id
           AND n.eleve_id = e.id
           AND n.subject_id = s.id
           AND n.trimestre = ?
        WHERE e.user_id = ? AND e.niveau = ?
        GROUP BY e.id
        ORDER BY moyenne_generale DESC, e.nom_complet COLLATE NOCASE ASC
        """,
        (int(trim), user_id, eleve["niveau"]),
    ).fetchall()

    scores = [(int(r["id"]), float(r["moyenne_generale"] or 0)) for r in class_rows]
    rank = next((i + 1 for i, item in enumerate(scores) if item[0] == eleve_id), 1)
    moyenne_classe = round(
        sum(item[1] for item in scores) / len(scores), 2
    ) if scores else 0.0

    return {
        "eleve": eleve,
        "subject_lines": subject_lines,
        "moyenne_generale": moyenne_generale,
        "moyenne_classe": moyenne_classe,
        "rank": rank,
        "total_eleves": len(scores),
    }


@bp.route("/sauvegarder_tout", methods=["POST"])
@login_required
@write_required
def sauvegarder_tout():
    user_id = session["user_id"]
    trim = _parse_trim(request.form.get("trimestre_save"))
    db = get_db()
    subjects = _get_subjects(db, user_id)
    subject_id = _select_subject_id(subjects, request.form.get("subject"))

    ids = request.form.getlist("id_eleve")
    devs = request.form.getlist("devoir")
    acts = request.form.getlist("activite")
    comps = request.form.getlist("compo")
    participations = request.form.getlist("participation")
    comportements = request.form.getlist("comportement")
    cahiers = request.form.getlist("cahier")
    projets = request.form.getlist("projet")
    assiduites = request.form.getlist("assiduite_outils")
    use_components = any([participations, comportements, cahiers, projets, assiduites])

    updated = 0
    for i in range(len(ids)):
        try:
            d = clean_note(_safe_list_get(devs, i))
            c = clean_note(_safe_list_get(comps, i))
            if use_components:
                p = _clean_component(_safe_list_get(participations, i), 3)
                b = _clean_component(_safe_list_get(comportements, i), 6)
                k = _clean_component(_safe_list_get(cahiers, i), 5)
                pr = _clean_component(_safe_list_get(projets, i), 4)
                ao = _clean_component(_safe_list_get(assiduites, i), 2)
            else:
                p, b, k, pr, ao = _split_activite_components(_safe_list_get(acts, i))
            a = _sum_activite_components(p, b, k, pr, ao)
            moy = ((d + a) / 2 + (c * 2)) / 3
            rem = get_appreciation_dynamique(moy, user_id)
            db.execute(
                """
                INSERT INTO notes (
                    user_id, eleve_id, subject_id, trimestre,
                    participation, comportement, cahier, projet, assiduite_outils,
                    activite, devoir, compo, remarques
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(user_id, eleve_id, subject_id, trimestre)
                DO UPDATE SET
                    participation=excluded.participation,
                    comportement=excluded.comportement,
                    cahier=excluded.cahier,
                    projet=excluded.projet,
                    assiduite_outils=excluded.assiduite_outils,
                    activite=excluded.activite,
                    devoir=excluded.devoir,
                    compo=excluded.compo,
                    remarques=excluded.remarques
                """,
                (user_id, ids[i], subject_id, int(trim), p, b, k, pr, ao, a, d, c, rem),
            )
            updated += 1
        except Exception:
            continue
    db.commit()
    log_change("update_notes", user_id, details=f"{updated} lignes", subject_id=subject_id)
    flash("Notes enregistrees.", "success")
    return redirect(request.referrer or url_for("main.index", trimestre=trim))


@bp.route("/ajouter_eleve", methods=["POST"])
@login_required
@write_required
def ajouter_eleve():
    user_id = session["user_id"]
    trim = _parse_trim(request.form.get("trimestre_ajout", "1"))
    d = clean_note(request.form.get("devoir"))
    c = clean_note(request.form.get("compo"))
    p_raw = (request.form.get("participation") or "").strip()
    b_raw = (request.form.get("comportement") or "").strip()
    k_raw = (request.form.get("cahier") or "").strip()
    pr_raw = (request.form.get("projet") or "").strip()
    ao_raw = (request.form.get("assiduite_outils") or "").strip()
    if any([p_raw, b_raw, k_raw, pr_raw, ao_raw]):
        p = _clean_component(p_raw, 3)
        b = _clean_component(b_raw, 6)
        k = _clean_component(k_raw, 5)
        pr = _clean_component(pr_raw, 4)
        ao = _clean_component(ao_raw, 2)
    else:
        p, b, k, pr, ao = _split_activite_components(request.form.get("activite"))
    a = _sum_activite_components(p, b, k, pr, ao)
    moy = ((d + a) / 2 + (c * 2)) / 3

    parent_phone = (request.form.get("parent_phone") or "").strip()
    parent_email = (request.form.get("parent_email") or "").strip()

    db = get_db()
    subjects = _get_subjects(db, user_id)
    subject_id = _select_subject_id(subjects, request.form.get("subject"))

    cur = db.execute(
        f"INSERT INTO eleves (user_id, nom_complet, niveau, remarques_t{trim}, devoir_t{trim}, activite_t{trim}, compo_t{trim}, parent_phone, parent_email) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (
            user_id,
            request.form["nom_complet"],
            request.form["niveau"],
            get_appreciation_dynamique(moy, user_id),
            d,
            a,
            c,
            parent_phone,
            parent_email,
        ),
    )
    eleve_id = cur.lastrowid

    db.execute(
        """
        INSERT INTO notes (
            user_id, eleve_id, subject_id, trimestre,
            participation, comportement, cahier, projet, assiduite_outils,
            activite, devoir, compo, remarques
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(user_id, eleve_id, subject_id, trimestre)
        DO UPDATE SET
            participation=excluded.participation,
            comportement=excluded.comportement,
            cahier=excluded.cahier,
            projet=excluded.projet,
            assiduite_outils=excluded.assiduite_outils,
            activite=excluded.activite,
            devoir=excluded.devoir,
            compo=excluded.compo,
            remarques=excluded.remarques
        """,
        (user_id, eleve_id, subject_id, int(trim), p, b, k, pr, ao, a, d, c, get_appreciation_dynamique(moy, user_id)),
    )

    db.commit()
    log_change("add_student", user_id, details=request.form.get("nom_complet", ""), eleve_id=eleve_id, subject_id=subject_id)
    return redirect(request.referrer or url_for("main.index", trimestre=trim))


@bp.route("/supprimer_multi", methods=["POST"])
@login_required
@write_required
def supprimer_multi():
    ids = request.form.getlist("ids")
    if ids:
        db = get_db()
        db.execute(
            f"DELETE FROM notes WHERE eleve_id IN ({','.join('?'*len(ids))}) AND user_id = ?",
            ids + [session["user_id"]],
        )
        db.execute(
            f"DELETE FROM eleves WHERE id IN ({','.join('?'*len(ids))}) AND user_id = ?",
            ids + [session["user_id"]],
        )
        db.commit()
        log_change("delete_students", session["user_id"], details=f"{len(ids)} eleves")
        flash(f"Supprimes ({len(ids)})", "success")
    return redirect(request.referrer or url_for("main.index"))


@bp.route("/import_excel", methods=["POST"])
@login_required
@write_required
def import_excel():
    user_id = session["user_id"]
    trim = _parse_trim(request.form.get("trimestre_import", "1"))
    file = request.files.get("fichier_excel")
    if not file or not file.filename:
        flash("Fichier Excel manquant.", "warning")
        return redirect(request.referrer or url_for("main.index", trimestre=trim))

    db = get_db()
    subjects = _get_subjects(db, user_id)
    subject_id = _select_subject_id(subjects, request.form.get("subject"))
    previous_meta = session.get("import_preview")
    if isinstance(previous_meta, dict):
        _clear_preview_meta(previous_meta)
    _cleanup_import_previews()

    token = uuid.uuid4().hex
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in (".xlsx", ".xls", ".xlsm"):
        ext = ".xlsx"
    preview_path = os.path.join(_preview_dir(), f"{token}{ext}")

    try:
        file.save(preview_path)
        all_sheets = pd.read_excel(preview_path, sheet_name=None, header=None)
    except Exception as exc:
        try:
            if os.path.exists(preview_path):
                os.remove(preview_path)
        except Exception:
            pass
        flash(f"Erreur lecture Excel: {exc}", "danger")
        return redirect(request.referrer or url_for("main.index", trimestre=trim))

    selected_sheet = ""
    selected_df = None
    header_detected = False
    for sheet_name, raw_df in (all_sheets or {}).items():
        prepared, _, detected = _prepare_import_dataframe(raw_df)
        if prepared is None or prepared.empty:
            continue
        selected_sheet = str(sheet_name)
        selected_df = prepared
        header_detected = detected
        break

    if selected_df is None or selected_df.empty:
        try:
            if os.path.exists(preview_path):
                os.remove(preview_path)
        except Exception:
            pass
        flash("Aucune ligne exploitable detectee dans le fichier.", "warning")
        return redirect(request.referrer or url_for("main.index", trimestre=trim))

    columns = [str(c) for c in selected_df.columns]
    defaults = _build_default_mapping(columns)
    sample_df = selected_df.head(8).copy()
    sample_rows = []
    for _, row in sample_df.iterrows():
        current = {}
        for col in columns:
            value = row.get(col, "")
            if pd.isna(value):
                value = ""
            current[col] = str(value).strip()
        sample_rows.append(current)

    session["import_preview"] = {
        "token": token,
        "path": preview_path,
        "trim": trim,
        "subject_id": int(subject_id),
        "sheet_name": selected_sheet,
        "created_at": int(datetime.now().timestamp()),
    }

    if not header_detected:
        flash("Entete non detectee automatiquement: verifiez bien la correspondance des colonnes.", "warning")

    return render_template(
        "import_mapping.html",
        token=token,
        mapping_fields=IMPORT_MAPPING_FIELDS,
        columns=columns,
        defaults=defaults,
        sample_rows=sample_rows,
        sample_headers=columns,
        source_sheet=selected_sheet,
        trim=trim,
        subject_id=subject_id,
    )


@bp.route("/import_excel_apply", methods=["POST"])
@login_required
@write_required
def import_excel_apply():
    user_id = session["user_id"]
    token = (request.form.get("token") or "").strip()
    meta = _get_preview_meta(token)
    if not meta:
        flash("Session d'import expiree. Recommencez l'import.", "warning")
        return redirect(url_for("main.index"))

    trim = _parse_trim(meta.get("trim"), "1")
    db = get_db()
    subjects = _get_subjects(db, user_id)
    subject_ids = {int(s["id"]) for s in subjects}
    try:
        subject_id = int(meta.get("subject_id"))
    except Exception:
        subject_id = None
    if subject_id not in subject_ids:
        subject_id = _select_subject_id(subjects, request.form.get("subject"))

    mapping = {}
    for key, _label in IMPORT_MAPPING_FIELDS:
        mapping[key] = (request.form.get(f"map_{key}") or "").strip()

    try:
        all_sheets = pd.read_excel(meta["path"], sheet_name=None, header=None)
    except Exception as exc:
        _clear_preview_meta(meta)
        flash(f"Lecture impossible pendant validation: {exc}", "danger")
        return redirect(url_for("main.index", trimestre=trim, subject=subject_id))

    inserted = 0
    updated = 0
    skipped_sheets = 0
    skipped_rows = 0
    use_components = any(
        mapping.get(k)
        for k in ("participation", "comportement", "cahier", "projet", "assiduite_outils")
    )

    for sheet_name, raw_df in (all_sheets or {}).items():
        prepared, _, _ = _prepare_import_dataframe(raw_df)
        if prepared is None or prepared.empty:
            skipped_sheets += 1
            continue

        columns = list(prepared.columns)
        resolved = {k: _resolve_mapped_column(columns, v) for k, v in mapping.items()}
        if not resolved.get("full_name") and not (
            resolved.get("last_name") or resolved.get("first_name")
        ):
            skipped_sheets += 1
            continue

        for _, row in prepared.iterrows():
            try:
                if resolved.get("full_name"):
                    full = str(_row_value(row, resolved["full_name"]) or "").strip()
                else:
                    last_name = str(_row_value(row, resolved.get("last_name")) or "").strip()
                    first_name = str(_row_value(row, resolved.get("first_name")) or "").strip()
                    full = f"{last_name} {first_name}".strip()

                if not full:
                    skipped_rows += 1
                    continue

                niveau = str(_row_value(row, resolved.get("classe")) or "").strip()
                if not niveau:
                    niveau = str(sheet_name).strip() or "Global"

                phone = str(_row_value(row, resolved.get("phone")) or "").strip()
                email = str(_row_value(row, resolved.get("email")) or "").strip()

                d = clean_note(_row_value(row, resolved.get("devoir")))
                c = clean_note(_row_value(row, resolved.get("compo")))

                if use_components:
                    p = _clean_component(_row_value(row, resolved.get("participation")), 3)
                    b = _clean_component(_row_value(row, resolved.get("comportement")), 6)
                    k = _clean_component(_row_value(row, resolved.get("cahier")), 5)
                    pr = _clean_component(_row_value(row, resolved.get("projet")), 4)
                    ao = _clean_component(_row_value(row, resolved.get("assiduite_outils")), 2)
                else:
                    p, b, k, pr, ao = _split_activite_components(
                        _row_value(row, resolved.get("activite"))
                    )
                a = _sum_activite_components(p, b, k, pr, ao)

                moy = ((d + a) / 2 + (c * 2)) / 3
                rem = get_appreciation_dynamique(moy, user_id)
                custom_rem = _row_value(row, resolved.get("remarques"))
                if custom_rem is not None and str(custom_rem).strip():
                    rem = str(custom_rem).strip()

                ex = db.execute(
                    "SELECT id FROM eleves WHERE nom_complet = ? AND niveau = ? AND user_id = ?",
                    (full, niveau, user_id),
                ).fetchone()

                if ex:
                    db.execute(
                        "UPDATE eleves SET parent_phone = COALESCE(?, parent_phone), parent_email = COALESCE(?, parent_email) WHERE id = ?",
                        (phone or None, email or None, ex["id"]),
                    )
                    db.execute(
                        """
                        INSERT INTO notes (
                            user_id, eleve_id, subject_id, trimestre,
                            participation, comportement, cahier, projet, assiduite_outils,
                            activite, devoir, compo, remarques
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(user_id, eleve_id, subject_id, trimestre)
                        DO UPDATE SET
                            participation=excluded.participation,
                            comportement=excluded.comportement,
                            cahier=excluded.cahier,
                            projet=excluded.projet,
                            assiduite_outils=excluded.assiduite_outils,
                            activite=excluded.activite,
                            devoir=excluded.devoir,
                            compo=excluded.compo,
                            remarques=excluded.remarques
                        """,
                        (user_id, ex["id"], subject_id, int(trim), p, b, k, pr, ao, a, d, c, rem),
                    )
                    updated += 1
                else:
                    cur = db.execute(
                        f"INSERT INTO eleves (user_id, nom_complet, niveau, remarques_t{trim}, devoir_t{trim}, activite_t{trim}, compo_t{trim}, parent_phone, parent_email) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (user_id, full, niveau, rem, d, a, c, phone, email),
                    )
                    eleve_id = cur.lastrowid
                    db.execute(
                        """
                        INSERT INTO notes (
                            user_id, eleve_id, subject_id, trimestre,
                            participation, comportement, cahier, projet, assiduite_outils,
                            activite, devoir, compo, remarques
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(user_id, eleve_id, subject_id, trimestre)
                        DO UPDATE SET
                            participation=excluded.participation,
                            comportement=excluded.comportement,
                            cahier=excluded.cahier,
                            projet=excluded.projet,
                            assiduite_outils=excluded.assiduite_outils,
                            activite=excluded.activite,
                            devoir=excluded.devoir,
                            compo=excluded.compo,
                            remarques=excluded.remarques
                        """,
                        (user_id, eleve_id, subject_id, int(trim), p, b, k, pr, ao, a, d, c, rem),
                    )
                    inserted += 1
            except Exception:
                skipped_rows += 1
                continue

    db.commit()
    _clear_preview_meta(meta)

    total = inserted + updated
    log_change(
        "import_excel",
        user_id,
        details=f"{total} lignes (new {inserted}, upd {updated}, sheets {skipped_sheets}, rows {skipped_rows})",
        subject_id=subject_id,
    )
    flash(
        f"Import termine: {total} lignes (nouveaux {inserted}, maj {updated}, onglets ignores {skipped_sheets}, lignes ignorees {skipped_rows})",
        "success",
    )
    return redirect(url_for("main.index", trimestre=trim, subject=subject_id))


@bp.route("/import_excel_cancel/<token>")
@login_required
def import_excel_cancel(token: str):
    meta = _get_preview_meta((token or "").strip())
    if meta:
        _clear_preview_meta(meta)
    return redirect(url_for("main.index"))

@bp.route("/remplir_bulletin_officiel", methods=["POST"])
@login_required
@write_required
def remplir_bulletin_officiel():
    user_id = session["user_id"]
    trim = _parse_trim(request.form.get("trimestre_fill", "1"))
    file = request.files.get("fichier_vide")
    if file and file.filename:
        try:
            wb = openpyxl.load_workbook(file)
            db = get_db()
            for sheet in wb.worksheets:
                header_row = None
                col_map = {}
                for i, row in enumerate(
                    sheet.iter_rows(min_row=1, max_row=20, values_only=True)
                ):
                    row_str = [str(c).lower() for c in row if c]
                    if any(x in row_str for x in ["nom", "Ø§Ù„Ù„Ù‚Ø¨"]):
                        header_row = i + 1
                        for cell in sheet[header_row]:
                            if not cell.value:
                                continue
                            v = str(cell.value).strip().lower()
                            if v in ["nom", "Ø§Ù„Ù„Ù‚Ø¨"]:
                                col_map["nom"] = cell.column
                            elif v in ["prenom", "Ø§Ù„Ø§Ø³Ù…"]:
                                col_map["prenom"] = cell.column
                            elif v in ["01", "act", "Ø§Ù„Ù†Ø´Ø§Ø·Ø§Øª"]:
                                col_map["act"] = cell.column
                            elif v in ["04", "dev", "Ø§Ù„ÙØ±Ø¶"]:
                                col_map["dev"] = cell.column
                            elif v in ["09", "compo", "Ø§Ù„Ø§Ø®ØªØ¨Ø§Ø±"]:
                                col_map["compo"] = cell.column
                            elif v in ["obs", "rem", "Ø§Ù„ØªÙ‚Ø¯ÙŠØ±Ø§Øª"]:
                                col_map["rem"] = cell.column
                        break

                if header_row and "nom" in col_map:
                    for r in range(header_row + 1, sheet.max_row + 1):
                        nom = sheet.cell(row=r, column=col_map["nom"]).value
                        if not nom:
                            continue
                        prenom = (
                            sheet.cell(row=r, column=col_map.get("prenom")).value
                            if col_map.get("prenom")
                            else ""
                        )
                        full = f"{nom} {prenom}".strip()
                        el = db.execute(
                            "SELECT * FROM eleves WHERE nom_complet = ? AND user_id = ?",
                            (full, user_id),
                        ).fetchone()
                        if el:
                            if "act" in col_map:
                                sheet.cell(row=r, column=col_map["act"]).value = el[
                                    f"activite_t{trim}"
                                ]
                            if "dev" in col_map:
                                sheet.cell(row=r, column=col_map["dev"]).value = el[
                                    f"devoir_t{trim}"
                                ]
                            if "compo" in col_map:
                                sheet.cell(row=r, column=col_map["compo"]).value = el[
                                    f"compo_t{trim}"
                                ]
                            if "rem" in col_map:
                                sheet.cell(row=r, column=col_map["rem"]).value = el[
                                    f"remarques_t{trim}"
                                ]
            out = BytesIO()
            wb.save(out)
            out.seek(0)
            return send_file(
                out,
                download_name="Bulletin_Rempli.xlsx",
                as_attachment=True,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            flash(f"Erreur: {e}", "danger")
    return redirect(request.referrer or url_for("main.index"))


@bp.route("/")
@login_required
def index():
    user_id = session["user_id"]
    trim = request.args.get("trimestre", "1")
    if trim not in ("1", "2", "3"):
        trim = "1"

    db = get_db()
    subjects = _get_subjects(db, user_id)
    subject_id = _select_subject_id(subjects, request.args.get("subject"))
    subject_name = next(s["name"] for s in subjects if int(s["id"]) == subject_id)
    role = session.get("role") or ("admin" if session.get("is_admin") else "prof")
    can_edit = role != "read_only"

    devoir_expr, activite_expr, compo_expr, remarques_expr, moy_expr = _note_expr(trim)
    filters = _build_filters(user_id, trim, request.args, moy_expr)
    niveau = filters["niveau"]
    search = filters["search"]
    sort = filters["sort"]
    order = filters["order"]
    etat = filters["etat"]
    min_moy = filters["min_moy"]
    max_moy = filters["max_moy"]

    try:
        page = int(request.args.get("page", "1") or 1)
    except ValueError:
        page = 1
    try:
        per_page = int(request.args.get("per_page", "50") or 50)
    except ValueError:
        per_page = 50

    page = max(1, page)
    per_page = min(200, max(10, per_page))

    classes = [
        r["niveau"]
        for r in db.execute(
            "SELECT DISTINCT niveau FROM eleves WHERE user_id = ? ORDER BY niveau",
            (user_id,),
        ).fetchall()
    ]

    where = filters["where"]
    params = filters["params"]

    join_params = [subject_id, int(trim)]

    stats_row = db.execute(
        f"""
        SELECT
          COUNT(*) AS nb_total,
          SUM(CASE WHEN {moy_expr} > 0 THEN 1 ELSE 0 END) AS nb_saisis,
          SUM(CASE WHEN {moy_expr} >= 10 THEN 1 ELSE 0 END) AS nb_admis,
          AVG(CASE WHEN {moy_expr} > 0 THEN {moy_expr} END) AS moyenne_generale,
          MAX(CASE WHEN {moy_expr} > 0 THEN {moy_expr} END) AS meilleure_note,
          MIN(CASE WHEN {moy_expr} > 0 THEN {moy_expr} END) AS pire_note
        FROM eleves e
        LEFT JOIN notes n ON n.user_id = e.user_id AND n.eleve_id = e.id AND n.subject_id = ? AND n.trimestre = ?
        WHERE {where}
        """,
        join_params + params,
    ).fetchone()

    total = int(stats_row["nb_total"] or 0)
    pages = max(1, (total + per_page - 1) // per_page)
    if page > pages:
        page = pages
    offset = (page - 1) * per_page

    direction = "DESC" if order == "desc" else "ASC"
    if sort == "name":
        order_clause = f"e.nom_complet COLLATE NOCASE {direction}, e.id ASC"
    elif sort == "moy":
        order_clause = f"moyenne {direction}, e.nom_complet COLLATE NOCASE ASC, e.id ASC"
    elif sort == "id":
        order_clause = f"e.id {direction}"
    else:
        order_clause = f"e.niveau COLLATE NOCASE {direction}, e.id ASC"

    rows = db.execute(
        f"""
        SELECT
          e.id,
          e.nom_complet,
          e.niveau,
          COALESCE(n.participation, 0) AS participation,
          COALESCE(n.comportement, 0) AS comportement,
          COALESCE(n.cahier, 0) AS cahier,
          COALESCE(n.projet, 0) AS projet,
          COALESCE(n.assiduite_outils, 0) AS assiduite_outils,
          {devoir_expr} AS devoir,
          {activite_expr} AS activite,
          {compo_expr} AS compo,
          {remarques_expr} AS remarques,
          ROUND({moy_expr}, 2) AS moyenne
        FROM eleves e
        LEFT JOIN notes n ON n.user_id = e.user_id AND n.eleve_id = e.id AND n.subject_id = ? AND n.trimestre = ?
        WHERE {where}
        ORDER BY {order_clause}
        LIMIT ? OFFSET ?
        """,
        join_params + params + [per_page, offset],
    ).fetchall()

    eleves_list = []
    for r in rows:
        p = float(r["participation"] or 0)
        b = float(r["comportement"] or 0)
        k = float(r["cahier"] or 0)
        pr = float(r["projet"] or 0)
        ao = float(r["assiduite_outils"] or 0)
        activite_value = float(r["activite"] or 0)
        if activite_value > 0 and p == 0 and b == 0 and k == 0 and pr == 0 and ao == 0:
            p, b, k, pr, ao = _split_activite_components(activite_value)

        eleves_list.append(
            {
                "id": r["id"],
                "nom_complet": r["nom_complet"],
                "niveau": r["niveau"],
                "remarques": r["remarques"],
                "devoir": float(r["devoir"] or 0),
                "activite": activite_value,
                "compo": float(r["compo"] or 0),
                "participation": p,
                "comportement": b,
                "cahier": k,
                "projet": pr,
                "assiduite_outils": ao,
                "moyenne": float(r["moyenne"] or 0),
            }
        )

    nb_admis = int(stats_row["nb_admis"] or 0)
    nb_total = total

    stats = {
        "moyenne_generale": round(float(stats_row["moyenne_generale"] or 0), 2),
        "meilleure_note": round(float(stats_row["meilleure_note"] or 0), 2),
        "pire_note": round(float(stats_row["pire_note"] or 0), 2),
        "nb_admis": nb_admis,
        "taux_reussite": round((nb_admis / nb_total) * 100, 1) if nb_total else 0,
        "nb_total": nb_total,
        "nb_saisis": int(stats_row["nb_saisis"] or 0),
    }

    class_rows = db.execute(
        f"""
        SELECT
          e.niveau,
          COUNT(*) AS total,
          AVG(CASE WHEN {moy_expr} > 0 THEN {moy_expr} END) AS avg_moy
        FROM eleves e
        LEFT JOIN notes n ON n.user_id = e.user_id AND n.eleve_id = e.id AND n.subject_id = ? AND n.trimestre = ?
        WHERE {where}
        GROUP BY e.niveau
        ORDER BY avg_moy DESC, e.niveau ASC
        """,
        join_params + params,
    ).fetchall()

    class_labels = [str(r["niveau"]) for r in class_rows]
    class_avgs = [round(float(r["avg_moy"] or 0), 2) for r in class_rows]

    dist_row = db.execute(
        f"""
        SELECT
          SUM(CASE WHEN {moy_expr} >= 10 THEN 1 ELSE 0 END) AS admis,
          SUM(CASE WHEN {moy_expr} > 0 AND {moy_expr} < 10 THEN 1 ELSE 0 END) AS echec,
          SUM(CASE WHEN {moy_expr} <= 0 THEN 1 ELSE 0 END) AS non_saisi
        FROM eleves e
        LEFT JOIN notes n ON n.user_id = e.user_id AND n.eleve_id = e.id AND n.subject_id = ? AND n.trimestre = ?
        WHERE {where}
        """,
        join_params + params,
    ).fetchone()

    dist_values = [
        int(dist_row["admis"] or 0),
        int(dist_row["echec"] or 0),
        int(dist_row["non_saisi"] or 0),
    ]

    top_rows = db.execute(
        f"""
        SELECT
          e.nom_complet,
          e.niveau,
          ROUND({moy_expr}, 2) AS moyenne
        FROM eleves e
        LEFT JOIN notes n ON n.user_id = e.user_id AND n.eleve_id = e.id AND n.subject_id = ? AND n.trimestre = ?
        WHERE {where}
        ORDER BY moyenne DESC, e.nom_complet ASC
        LIMIT 10
        """,
        join_params + params,
    ).fetchall()

    top_eleves = [
        {"nom": r["nom_complet"], "niveau": r["niveau"], "moyenne": float(r["moyenne"] or 0)}
        for r in top_rows
    ]

    risk_where = "e.user_id = ?"
    risk_params = [user_id]
    if niveau and niveau != "all":
        risk_where += " AND e.niveau = ?"
        risk_params.append(niveau)
    if search:
        risk_where += " AND e.nom_complet LIKE ?"
        risk_params.append(f"%{search}%")

    risk_count_row = db.execute(
        f"""
        SELECT COUNT(*) AS c
        FROM eleves e
        LEFT JOIN notes n ON n.user_id = e.user_id AND n.eleve_id = e.id AND n.subject_id = ? AND n.trimestre = ?
        WHERE {risk_where} AND {moy_expr} > 0 AND {moy_expr} < 10
        """,
        join_params + risk_params,
    ).fetchone()
    risk_count = int(risk_count_row["c"] or 0)

    risk_rows = db.execute(
        f"""
        SELECT
          e.id,
          e.nom_complet,
          e.niveau,
          ROUND({moy_expr}, 2) AS moyenne
        FROM eleves e
        LEFT JOIN notes n ON n.user_id = e.user_id AND n.eleve_id = e.id AND n.subject_id = ? AND n.trimestre = ?
        WHERE {risk_where} AND {moy_expr} > 0 AND {moy_expr} < 10
        ORDER BY moyenne ASC, e.nom_complet ASC
        LIMIT 8
        """,
        join_params + risk_params,
    ).fetchall()
    risk_students = [
        {
            "id": int(r["id"]),
            "nom": r["nom_complet"],
            "niveau": r["niveau"],
            "moyenne": float(r["moyenne"] or 0),
        }
        for r in risk_rows
    ]

    progress_where = "e.user_id = ?"
    progress_params = [user_id]
    if niveau and niveau != "all":
        progress_where += " AND e.niveau = ?"
        progress_params.append(niveau)
    if search:
        progress_where += " AND e.nom_complet LIKE ?"
        progress_params.append(f"%{search}%")

    m1 = "((COALESCE(n1.devoir, e.devoir_t1) + COALESCE(n1.activite, e.activite_t1))/2.0 + (COALESCE(n1.compo, e.compo_t1) * 2.0))/3.0"
    m2 = "((COALESCE(n2.devoir, e.devoir_t2) + COALESCE(n2.activite, e.activite_t2))/2.0 + (COALESCE(n2.compo, e.compo_t2) * 2.0))/3.0"
    m3 = "((COALESCE(n3.devoir, e.devoir_t3) + COALESCE(n3.activite, e.activite_t3))/2.0 + (COALESCE(n3.compo, e.compo_t3) * 2.0))/3.0"

    progress_rows = db.execute(
        f"""
        SELECT
          e.niveau,
          AVG(CASE WHEN {m1} > 0 THEN {m1} END) AS t1,
          AVG(CASE WHEN {m2} > 0 THEN {m2} END) AS t2,
          AVG(CASE WHEN {m3} > 0 THEN {m3} END) AS t3
        FROM eleves e
        LEFT JOIN notes n1 ON n1.user_id = e.user_id AND n1.eleve_id = e.id AND n1.subject_id = ? AND n1.trimestre = 1
        LEFT JOIN notes n2 ON n2.user_id = e.user_id AND n2.eleve_id = e.id AND n2.subject_id = ? AND n2.trimestre = 2
        LEFT JOIN notes n3 ON n3.user_id = e.user_id AND n3.eleve_id = e.id AND n3.subject_id = ? AND n3.trimestre = 3
        WHERE {progress_where}
        GROUP BY e.niveau
        ORDER BY e.niveau COLLATE NOCASE ASC
        """,
        [subject_id, subject_id, subject_id] + progress_params,
    ).fetchall()

    progress_labels = [str(r["niveau"]) for r in progress_rows]
    progress_t1 = [round(float(r["t1"] or 0), 2) for r in progress_rows]
    progress_t2 = [round(float(r["t2"] or 0), 2) for r in progress_rows]
    progress_t3 = [round(float(r["t3"] or 0), 2) for r in progress_rows]

    chart_data = {
        "classes": {"labels": class_labels, "values": class_avgs},
        "distribution": {"labels": ["Admis", "Echec", "Non saisi"], "values": dist_values},
        "progression": {
            "labels": progress_labels,
            "t1": progress_t1,
            "t2": progress_t2,
            "t3": progress_t3,
        },
    }

    base_args = dict(request.args)
    base_args.pop("page", None)
    risk_args = dict(request.args)
    risk_args["etat"] = "echec"
    risk_args.pop("page", None)
    risk_url = url_for("main.index", **risk_args)

    return render_template(
        "index.html",
        eleves=eleves_list,
        stats=stats,
        trimestre=trim,
        nom_prof=session.get("nom_affichage"),
        niveau_actuel=niveau,
        recherche_actuelle=search,
        liste_classes=classes,
        sort=sort,
        order=order,
        page=page,
        pages=pages,
        per_page=per_page,
        total=total,
        base_args=base_args,
        min_moy=min_moy,
        max_moy=max_moy,
        etat=etat,
        chart_data=chart_data,
        top_eleves=top_eleves,
        risk_students=risk_students,
        risk_count=risk_count,
        risk_url=risk_url,
        subjects=subjects,
        subject_id=subject_id,
        subject_name=subject_name,
        can_edit=can_edit,
        school_year=_school_year(datetime.now()),
        school_name=session.get("school_name"),
    )


@bp.route("/bulletin/<int:id>")
@login_required
def bulletin(id: int):
    user_id = session["user_id"]
    trim = _parse_trim(request.args.get("trimestre", "1"))

    db = get_db()
    data = _build_bulletin_multisubject(db, user_id, id, trim)
    if not data:
        return "Eleve introuvable"

    return render_template(
        "bulletin.html",
        eleve=data["eleve"],
        subject_lines=data["subject_lines"],
        rank=data["rank"],
        total_eleves=data["total_eleves"],
        moyenne_generale=data["moyenne_generale"],
        moyenne_classe=data["moyenne_classe"],
        trimestre=trim,
        nom_prof=session.get("nom_affichage"),
    )


@bp.route("/bulletin_pdf/<int:id>")
@login_required
def bulletin_pdf(id: int):
    user_id = session["user_id"]
    trim = _parse_trim(request.args.get("trimestre", "1"))

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except Exception:
        flash("PDF indisponible. Installez reportlab (pip install reportlab).", "danger")
        return redirect(url_for("main.bulletin", id=id, trimestre=trim))

    db = get_db()
    data = _build_bulletin_multisubject(db, user_id, id, trim)
    if not data:
        return "Eleve introuvable"
    eleve = data["eleve"]
    subject_lines = data["subject_lines"]
    moyenne_generale = data["moyenne_generale"]
    moyenne_classe = data["moyenne_classe"]
    rank = data["rank"]
    total_eleves = data["total_eleves"]

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )
    styles = getSampleStyleSheet()

    story = []
    school_name = session.get("school_name") or os.environ.get("SCHOOL_NAME", "Etablissement")
    school_year = _school_year(datetime.now())
    logo_path = os.path.join(BASE_DIR, "static", "logo.png")
    stamp_path = os.path.join(BASE_DIR, "static", "stamp.png")

    header_right = Paragraph(
        f"<b>{school_name}</b><br/>Bulletin de notes<br/>Annee {school_year}",
        styles["Heading2"],
    )
    header_left = ""
    if os.path.exists(logo_path):
        from reportlab.platypus import Image

        header_left = Image(logo_path, width=28 * mm, height=28 * mm)

    header_table = Table([[header_left, header_right]], colWidths=[32 * mm, 150 * mm])
    header_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (1, 0), (1, 0), "CENTER"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(header_table)
    story.append(Spacer(1, 6))

    info_data = [
        ["Eleve", eleve["nom_complet"]],
        ["Classe", eleve["niveau"]],
        ["Trimestre", trim],
        ["Nombre matieres", len(subject_lines)],
        ["Prof", session.get("nom_affichage", "")],
    ]
    info_table = Table(info_data, colWidths=[28 * mm, 120 * mm])
    info_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.3, colors.black),
                ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(info_table)
    story.append(Spacer(1, 10))

    table_data = [
        ["Matiere", "Activite", "Devoir", "Compo", "Moyenne", "Remarques"],
    ]
    for line in subject_lines:
        table_data.append(
            [
                line["subject_name"],
                line["activite"],
                line["devoir"],
                line["compo"],
                line["moyenne"],
                line["remarques"],
            ]
        )

    table = Table(
        table_data,
        colWidths=[35 * mm, 22 * mm, 22 * mm, 22 * mm, 22 * mm, 51 * mm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("ALIGN", (1, 1), (4, 1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 14))

    summary_data = [
        ["Moyenne classe", moyenne_classe],
        ["Rang", f"{rank} / {total_eleves}"],
        ["Moyenne generale", moyenne_generale],
    ]
    summary = Table(summary_data, colWidths=[60 * mm, 40 * mm])
    summary.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ]
        )
    )
    story.append(summary)
    story.append(Spacer(1, 12))
    story.append(Paragraph("Cachet et signature", styles["Normal"]))
    if os.path.exists(stamp_path):
        from reportlab.platypus import Image

        stamp = Image(stamp_path, width=28 * mm, height=28 * mm)
        stamp.hAlign = "RIGHT"
        story.append(stamp)

    doc.build(story)
    buffer.seek(0)

    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", eleve["nom_complet"])
    filename = f"bulletin_{safe_name}_T{trim}.pdf"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )


@bp.route("/print_list")
@login_required
def print_list():
    user_id = session["user_id"]
    trim = request.args.get("trimestre", "1")
    if trim not in ("1", "2", "3"):
        trim = "1"

    db = get_db()
    subjects = _get_subjects(db, user_id)
    subject_id = _select_subject_id(subjects, request.args.get("subject"))
    subject_name = next(s["name"] for s in subjects if int(s["id"]) == subject_id)

    devoir_expr, activite_expr, compo_expr, remarques_expr, moy_expr = _note_expr(trim)
    filters = _build_filters(user_id, trim, request.args, moy_expr)
    niveau = filters["niveau"]

    where = filters["where"]
    params = filters["params"]

    rows = db.execute(
        f"""
        SELECT
          e.nom_complet,
          e.niveau,
          {activite_expr} AS activite,
          {devoir_expr} AS devoir,
          {compo_expr} AS compo,
          ROUND({moy_expr}, 2) AS moyenne,
          {remarques_expr} AS remarques
        FROM eleves e
        LEFT JOIN notes n ON n.user_id = e.user_id AND n.eleve_id = e.id AND n.subject_id = ? AND n.trimestre = ?
        WHERE {where}
        ORDER BY e.niveau, e.id
        """,
        [subject_id, int(trim)] + params,
    ).fetchall()

    eleves_list = [
        {
            "nom_complet": r["nom_complet"],
            "niveau": r["niveau"],
            "activite": r["activite"],
            "devoir": r["devoir"],
            "compo": r["compo"],
            "moyenne": float(r["moyenne"] or 0),
            "remarques": r["remarques"],
        }
        for r in rows
    ]

    return render_template(
        "print_template.html",
        eleves=eleves_list,
        nom_prof=session.get("nom_affichage"),
        trimestre=trim,
        niveau=niveau,
        subject_name=subject_name,
    )


@bp.route("/export_list_pdf")
@login_required
def export_list_pdf():
    user_id = session["user_id"]
    trim = request.args.get("trimestre", "1")
    if trim not in ("1", "2", "3"):
        trim = "1"

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except Exception:
        flash("PDF indisponible. Installez reportlab (pip install reportlab).", "danger")
        return redirect(request.referrer or url_for("main.index", trimestre=trim))

    db = get_db()
    subjects = _get_subjects(db, user_id)
    subject_id = _select_subject_id(subjects, request.args.get("subject"))
    subject_name = next(s["name"] for s in subjects if int(s["id"]) == subject_id)

    devoir_expr, activite_expr, compo_expr, remarques_expr, moy_expr = _note_expr(trim)
    filters = _build_filters(user_id, trim, request.args, moy_expr)

    where = filters["where"]
    params = filters["params"]

    rows = db.execute(
        f"""
        SELECT
          e.nom_complet,
          e.niveau,
          {activite_expr} AS activite,
          {devoir_expr} AS devoir,
          {compo_expr} AS compo,
          ROUND({moy_expr}, 2) AS moyenne,
          {remarques_expr} AS remarques
        FROM eleves e
        LEFT JOIN notes n ON n.user_id = e.user_id AND n.eleve_id = e.id AND n.subject_id = ? AND n.trimestre = ?
        WHERE {where}
        ORDER BY e.niveau, e.id
        """,
        [subject_id, int(trim)] + params,
    ).fetchall()

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name = "Helvetica"
    font_bold = "Helvetica-Bold"
    font_path = os.path.join(BASE_DIR, "static", "fonts", "TimesNewRoman-Regular.ttf")
    font_bold_path = os.path.join(BASE_DIR, "static", "fonts", "TimesNewRoman-Bold.ttf")
    if os.path.exists(font_path) and os.path.exists(font_bold_path):
        try:
            pdfmetrics.registerFont(TTFont("Arabic", font_path))
            pdfmetrics.registerFont(TTFont("ArabicBold", font_bold_path))
            font_name = "Arabic"
            font_bold = "ArabicBold"
        except Exception:
            font_name = "Helvetica"
            font_bold = "Helvetica-Bold"

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()

    story = []
    school_name = session.get("school_name") or os.environ.get("SCHOOL_NAME", "")
    class_label = filters.get("niveau") if "filters" in locals() else ""
    class_suffix = f" - {class_label}" if class_label and class_label != "all" else ""
    title = _arabize(f"\u0642\u0627\u0626\u0645\u0629 \u0627\u0644\u062a\u0644\u0627\u0645\u064a\u0630{class_suffix} - T{trim} - {subject_name}")
    title_style = styles["Title"].clone("ArabicTitle")
    title_style.fontName = font_bold
    title_style.alignment = 1
    if school_name:
        school_style = styles["Normal"].clone("ArabicSchool")
        school_style.fontName = font_bold
        school_style.alignment = 1
        story.append(Paragraph(_arabize(school_name), school_style))
        story.append(Spacer(1, 4))
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 6))

    prof_name = session.get("nom_affichage")
    if prof_name:
        prof_style = styles["Normal"].clone("ArabicProf")
        prof_style.fontName = font_name
        prof_style.alignment = 1
        story.append(Paragraph(_arabize(f"Ø§Ù„Ø£Ø³ØªØ§Ø°: {prof_name}"), prof_style))
        story.append(Spacer(1, 6))

    table_data = [[
        _arabize("\u0627\u0644\u0631\u0642\u0645"),
        _arabize("\u0627\u0644\u0627\u0633\u0645 \u0648 \u0627\u0644\u0644\u0642\u0628"),
        _arabize("\u0627\u0644\u0642\u0633\u0645"),
        _arabize("\u0627\u0644\u0646\u0634\u0627\u0637"),
        _arabize("\u0627\u0644\u0641\u0631\u0636"),
        _arabize("\u0627\u0644\u0627\u062e\u062a\u0628\u0627\u0631"),
        _arabize("\u0627\u0644\u0645\u0639\u062f\u0644"),
        _arabize("\u0627\u0644\u0645\u0644\u0627\u062d\u0638\u0627\u062a"),
    ]]
    for i, r in enumerate(rows, 1):
        table_data.append(
            [
                i,
                _arabize(r["nom_complet"]),
                _arabize(r["niveau"]),
                r["activite"],
                r["devoir"],
                r["compo"],
                r["moyenne"],
                _arabize(r["remarques"] or ""),
            ]
        )

    # RTL: invert column order so first column is on the right
    table_data = [list(reversed(row)) for row in table_data]
    col_widths = [10 * mm, 55 * mm, 26 * mm, 22 * mm, 22 * mm, 22 * mm, 22 * mm, 80 * mm]
    col_widths = list(reversed(col_widths))

    table = Table(
        table_data,
        repeatRows=1,
        colWidths=col_widths,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTNAME", (0, 0), (-1, 0), font_bold),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("ALIGN", (0, 1), (0, -1), "RIGHT"),
                ("ALIGN", (1, 1), (1, -1), "CENTER"),
                ("ALIGN", (2, 1), (2, -1), "CENTER"),
                ("ALIGN", (3, 1), (3, -1), "CENTER"),
                ("ALIGN", (4, 1), (4, -1), "CENTER"),
                ("ALIGN", (5, 1), (5, -1), "CENTER"),
                ("ALIGN", (6, 1), (6, -1), "RIGHT"),
                ("ALIGN", (7, 1), (7, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(table)

    doc.build(story)
    buffer.seek(0)

    safe_subject = re.sub(r"[^A-Za-z0-9_-]+", "_", subject_name)
    filename = f"liste_eleves_T{trim}_{safe_subject}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")


@bp.route("/export_excel")
@login_required
def export_excel():
    user_id = session["user_id"]
    trim = request.args.get("trimestre", "1")
    if trim not in ("1", "2", "3"):
        trim = "1"

    db = get_db()
    subjects = _get_subjects(db, user_id)
    subject_id = _select_subject_id(subjects, request.args.get("subject"))

    devoir_expr, activite_expr, compo_expr, remarques_expr, moy_expr = _note_expr(trim)
    filters = _build_filters(user_id, trim, request.args, moy_expr)
    where = filters["where"]
    params = filters["params"]
    sort = filters["sort"]
    order = filters["order"]

    direction = "DESC" if order == "desc" else "ASC"
    if sort == "name":
        order_clause = f"e.nom_complet COLLATE NOCASE {direction}, e.id ASC"
    elif sort == "moy":
        order_clause = f"moyenne {direction}, e.nom_complet COLLATE NOCASE ASC, e.id ASC"
    elif sort == "id":
        order_clause = f"e.id {direction}"
    else:
        order_clause = f"e.niveau COLLATE NOCASE {direction}, e.id ASC"

    rows = db.execute(
        f"""
        SELECT
          e.id,
          e.nom_complet,
          e.niveau,
          {devoir_expr} AS devoir,
          {activite_expr} AS activite,
          {compo_expr} AS compo,
          {remarques_expr} AS remarques,
          ROUND({moy_expr}, 2) AS moyenne
        FROM eleves e
        LEFT JOIN notes n ON n.user_id = e.user_id AND n.eleve_id = e.id AND n.subject_id = ? AND n.trimestre = ?
        WHERE {where}
        ORDER BY {order_clause}
        """,
        [subject_id, int(trim)] + params,
    ).fetchall()

    data = []
    for r in rows:
        moy = float(r["moyenne"] or 0)
        if moy >= 10:
            etat = "Admis"
        elif moy > 0:
            etat = "Echec"
        else:
            etat = "Non saisi"
        data.append(
            {
                "ID": r["id"],
                "Nom complet": r["nom_complet"],
                "Classe": r["niveau"],
                "Activite": r["activite"],
                "Devoir": r["devoir"],
                "Compo": r["compo"],
                "Moyenne": moy,
                "Etat": etat,
                "Remarques": r["remarques"],
            }
        )

    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=f"T{trim}")
    output.seek(0)

    filename = f"export_eleves_T{trim}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/export_parents")
@login_required
def export_parents():
    user_id = session["user_id"]
    trim = request.args.get("trimestre", "1")
    if trim not in ("1", "2", "3"):
        trim = "1"

    db = get_db()
    subjects = _get_subjects(db, user_id)
    subject_id = _select_subject_id(subjects, request.args.get("subject"))

    devoir_expr, activite_expr, compo_expr, remarques_expr, moy_expr = _note_expr(trim)
    filters = _build_filters(user_id, trim, request.args, moy_expr)
    where = filters["where"]
    params = filters["params"]

    rows = db.execute(
        f"""
        SELECT
          e.nom_complet,
          e.niveau,
          e.parent_phone,
          e.parent_email,
          ROUND({moy_expr}, 2) AS moyenne,
          {remarques_expr} AS remarques
        FROM eleves e
        LEFT JOIN notes n ON n.user_id = e.user_id AND n.eleve_id = e.id AND n.subject_id = ? AND n.trimestre = ?
        WHERE {where}
        ORDER BY e.niveau, e.nom_complet
        """,
        [subject_id, int(trim)] + params,
    ).fetchall()

    data = []
    for r in rows:
        moy = float(r["moyenne"] or 0)
        if moy >= 10:
            etat = "Admis"
        elif moy > 0:
            etat = "Echec"
        else:
            etat = "Non saisi"
        data.append(
            {
                "Eleve": r["nom_complet"],
                "Classe": r["niveau"],
                "Tel parent": r["parent_phone"] or "",
                "Email parent": r["parent_email"] or "",
                "Moyenne": moy,
                "Etat": etat,
                "Remarques": r["remarques"],
                "Message": "",
            }
        )

    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=f"Parents_T{trim}")
    output.seek(0)

    filename = f"export_parents_T{trim}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/settings", methods=["GET", "POST"])
@login_required
@write_required
def settings():
    user_id = session["user_id"]
    db = get_db()

    if request.method == "POST":
        mins = request.form.getlist("min_val")
        maxs = request.form.getlist("max_val")
        msgs = request.form.getlist("message")

        db.execute("DELETE FROM appreciations WHERE user_id = ?", (user_id,))
        for i in range(len(mins)):
            if not mins[i]:
                continue
            try:
                min_v = float(mins[i])
                max_v = float(maxs[i]) if maxs[i] else min_v
            except ValueError:
                continue
            db.execute(
                "INSERT INTO appreciations (user_id, min_val, max_val, message) VALUES (?, ?, ?, ?)",
                (user_id, min_v, max_v, msgs[i]),
            )
        db.commit()

        # Recalcul rapide des remarques (notes)
        notes = db.execute(
            "SELECT id, activite, devoir, compo FROM notes WHERE user_id = ?",
            (user_id,),
        ).fetchall()
        for n in notes:
            moy = ((n["devoir"] + n["activite"]) / 2 + (n["compo"] * 2)) / 3
            db.execute(
                "UPDATE notes SET remarques = ? WHERE id = ?",
                (get_appreciation_dynamique(moy, user_id), n["id"]),
            )

        # Legacy recalcul (old columns)
        for el in db.execute(
            "SELECT * FROM eleves WHERE user_id = ?",
            (user_id,),
        ).fetchall():
            for t in range(1, 4):
                moy = ((el[f"devoir_t{t}"] + el[f"activite_t{t}"]) / 2 + (el[f"compo_t{t}"] * 2)) / 3
                db.execute(
                    f"UPDATE eleves SET remarques_t{t} = ? WHERE id = ?",
                    (get_appreciation_dynamique(moy, user_id), el["id"]),
                )
        db.commit()
        flash("Sauvegarde", "success")

    rules = db.execute(
        "SELECT * FROM appreciations WHERE user_id = ? ORDER BY min_val",
        (user_id,),
    ).fetchall()
    return render_template("settings.html", rules=rules)


@bp.route("/timetable", methods=["GET", "POST"])
@login_required
@write_required
def timetable():
    user_id = session["user_id"]
    db = get_db()

    classes = [
        r["niveau"]
        for r in db.execute(
            "SELECT DISTINCT niveau FROM eleves WHERE user_id = ? ORDER BY niveau",
            (user_id,),
        ).fetchall()
    ]

    niveau = request.values.get("niveau", "")
    days = [
        {"key": "sun", "label": "Dimanche"},
        {"key": "mon", "label": "Lundi"},
        {"key": "tue", "label": "Mardi"},
        {"key": "wed", "label": "Mercredi"},
        {"key": "thu", "label": "Jeudi"},
    ]
    slots = [
        {"key": "s1", "label": "08:00-09:00"},
        {"key": "s2", "label": "09:00-10:00"},
        {"key": "s3", "label": "10:00-11:00"},
        {"key": "s4", "label": "11:00-12:00"},
        {"key": "s5", "label": "12:00-13:00"},
        {"key": "s6", "label": "13:00-14:00"},
        {"key": "s7", "label": "14:00-15:00"},
        {"key": "s8", "label": "15:00-16:00"},
        {"key": "s9", "label": "16:00-17:00"},
    ]

    if request.method == "POST" and niveau:
        updated = 0
        for d in days:
            for s in slots:
                key = f"cell_{d['key']}_{s['key']}"
                value = (request.form.get(key) or "").strip()
                if value:
                    db.execute(
                        """
                        INSERT INTO timetable (user_id, niveau, day, slot, label)
                        VALUES (?, ?, ?, ?, ?)
                        ON CONFLICT(user_id, niveau, day, slot)
                        DO UPDATE SET label=excluded.label
                        """,
                        (user_id, niveau, d["key"], s["key"], value),
                    )
                    updated += 1
                else:
                    db.execute(
                        "DELETE FROM timetable WHERE user_id = ? AND niveau = ? AND day = ? AND slot = ?",
                        (user_id, niveau, d["key"], s["key"]),
                    )
        db.commit()
        log_change("update_timetable", user_id, details=f"{niveau} ({updated} cases)")
        flash("Emploi du temps enregistre.", "success")
        return redirect(url_for("main.timetable", niveau=niveau))

    grid = {}
    if niveau:
        rows = db.execute(
            "SELECT day, slot, label FROM timetable WHERE user_id = ? AND niveau = ?",
            (user_id, niveau),
        ).fetchall()
        for r in rows:
            grid.setdefault(r["day"], {})[r["slot"]] = r["label"]

    return render_template(
        "timetable.html",
        liste_classes=classes,
        niveau_actuel=niveau,
        days=days,
        slots=slots,
        grid=grid,
    )


@bp.route("/subjects", methods=["GET", "POST"])
@login_required
@write_required
def subjects():
    user_id = session["user_id"]
    db = get_db()
    if not session.get("is_admin"):
        try:
            row = db.execute(
                "SELECT lock_subject FROM users WHERE id = ?",
                (user_id,),
            ).fetchone()
            if row and int(row["lock_subject"] or 0) == 1:
                flash("Acces reserve a l'administration.", "warning")
                return redirect(url_for("main.index"))
        except Exception:
            pass
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        if not name:
            flash("Nom manquant", "danger")
        else:
            try:
                db.execute(
                    "INSERT INTO subjects (user_id, name) VALUES (?, ?)",
                    (user_id, name),
                )
                db.commit()
                log_change("add_subject", user_id, details=name)
                flash("Matiere ajoutee", "success")
            except Exception:
                flash("Matiere existe deja", "warning")

    subjects_list = db.execute(
        "SELECT id, name FROM subjects WHERE user_id = ? ORDER BY name",
        (user_id,),
    ).fetchall()
    return render_template("subjects.html", subjects=subjects_list)


@bp.route("/subjects/delete/<int:subject_id>", methods=["POST"])
@login_required
@write_required
def delete_subject(subject_id: int):
    user_id = session["user_id"]
    db = get_db()
    count = db.execute(
        "SELECT COUNT(*) as c FROM subjects WHERE user_id = ?",
        (user_id,),
    ).fetchone()["c"]
    if count <= 1:
        flash("Impossible de supprimer la derniere matiere", "warning")
        return redirect(url_for("main.subjects"))

    db.execute(
        "DELETE FROM notes WHERE user_id = ? AND subject_id = ?",
        (user_id, subject_id),
    )
    db.execute(
        "DELETE FROM subjects WHERE user_id = ? AND id = ?",
        (user_id, subject_id),
    )
    db.commit()
    log_change("delete_subject", user_id, details=str(subject_id), subject_id=subject_id)
    flash("Matiere supprimee", "success")
    return redirect(url_for("main.subjects"))


@bp.route("/history")
@login_required
def history():
    user_id = session["user_id"]
    db = get_db()
    actions = [
        r["action"]
        for r in db.execute(
            "SELECT DISTINCT action FROM change_log WHERE user_id = ? ORDER BY action",
            (user_id,),
        ).fetchall()
    ]
    subjects = _get_subjects(db, user_id)

    filters = _build_history_filters(user_id, request.args)
    where = filters["where"]
    params = filters["params"]

    rows = db.execute(
        f"""
        SELECT l.*, e.nom_complet AS eleve_name, s.name AS subject_name
        FROM change_log l
        LEFT JOIN eleves e ON e.id = l.eleve_id
        LEFT JOIN subjects s ON s.id = l.subject_id
        WHERE {where}
        ORDER BY l.created_at DESC
        LIMIT 200
        """,
        params,
    ).fetchall()

    logs = []
    for r in rows:
        item = dict(r)
        try:
            item["time"] = datetime.fromtimestamp(item["created_at"]).strftime("%Y-%m-%d %H:%M")
        except Exception:
            item["time"] = str(item.get("created_at", ""))
        logs.append(item)

    export_url = url_for("main.history_export", **dict(request.args))
    return render_template(
        "history.html",
        logs=logs,
        actions=actions,
        subjects=subjects,
        filters=filters,
        export_url=export_url,
        total=len(logs),
    )


@bp.route("/history/export")
@login_required
def history_export():
    user_id = session["user_id"]
    db = get_db()

    filters = _build_history_filters(user_id, request.args)
    where = filters["where"]
    params = filters["params"]

    rows = db.execute(
        f"""
        SELECT l.*, e.nom_complet AS eleve_name, s.name AS subject_name
        FROM change_log l
        LEFT JOIN eleves e ON e.id = l.eleve_id
        LEFT JOIN subjects s ON s.id = l.subject_id
        WHERE {where}
        ORDER BY l.created_at DESC
        """,
        params,
    ).fetchall()

    text_out = StringIO(newline="")
    writer = csv.writer(text_out)
    writer.writerow(["Date", "Action", "Eleve", "Matiere", "Details"])
    for r in rows:
        try:
            date_val = datetime.fromtimestamp(r["created_at"]).strftime("%Y-%m-%d %H:%M")
        except Exception:
            date_val = str(r.get("created_at", ""))
        writer.writerow(
            [
                date_val,
                r["action"],
                r["eleve_name"] or "",
                r["subject_name"] or "",
                r["details"] or "",
            ]
        )
    payload = text_out.getvalue().encode("utf-8-sig")
    output = BytesIO(payload)
    output.seek(0)

    filename = f"historique_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="text/csv")


