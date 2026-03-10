import os
import json
import uuid
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime
from flask import Blueprint, render_template, request, session, redirect, url_for, flash, send_file

from core.audit import log_change
from core.db import get_db
from core.security import login_required, write_required
from core.utils import clean_note, get_appreciation_dynamique
from edumaster.services.common import (
    get_subjects,
    get_user_assignment_scope,
    parse_trim,
    resolve_school_year,
    select_subject_id,
)
from edumaster.services.grading import clean_component, split_activite_components, sum_activite_components
from edumaster.services.import_utils import (
    preview_dir, cleanup_import_previews, get_preview_meta, clear_preview_meta,
    prepare_import_dataframe, build_default_mapping, resolve_mapped_column, row_value
)
from edumaster.services.scan_import import (
    ScanImportError,
    build_student_catalog,
    extract_rows_from_scanned_pdf,
    match_scanned_row,
)

bp = Blueprint("imports", __name__)

IMPORT_MAPPING_FIELDS = [
    ("full_name", "Nom complet"),
    ("last_name", "Nom"),
    ("first_name", "Prenom"),
    ("classe", "Classe"),
    ("devoir", "Devoir (/20)"),
    ("activite", "Activite (/20)"),
    ("compo", "Compo (/20)"),
    ("participation", "Participation (/3)"),
    ("comportement", "Comportement (/6)"),
    ("cahier", "Cahier (/5)"),
    ("projet", "Projet (/4)"),
    ("assiduite_outils", "Absences/Outils (/2)"),
    ("remarques", "Remarques"),
    ("phone", "Telephone parent"),
    ("email", "Email parent"),
]

SCAN_PREVIEW_SESSION_KEY = "import_scan_preview"


def _scan_preview_paths(token: str):
    folder = preview_dir()
    return (
        os.path.join(folder, f"{token}.pdf"),
        os.path.join(folder, f"{token}.scan.json"),
    )


def _get_scan_preview_meta(token: str):
    data = session.get(SCAN_PREVIEW_SESSION_KEY) or {}
    if data.get("token") != token:
        return None
    path = data.get("path") or ""
    json_path = data.get("json_path") or ""
    if not path or not os.path.exists(path) or not json_path or not os.path.exists(json_path):
        return None
    return data


def _clear_scan_preview_meta(meta):
    if isinstance(meta, dict):
        for key in ("path", "json_path"):
            target = meta.get(key) or ""
            try:
                if target and os.path.exists(target):
                    os.remove(target)
            except Exception:
                pass
    session.pop(SCAN_PREVIEW_SESSION_KEY, None)


def _save_scan_preview_rows(json_path: str, rows):
    with open(json_path, "w", encoding="utf-8") as handle:
        json.dump(rows, handle, ensure_ascii=False)


def _load_scan_preview_rows(meta):
    with open(meta["json_path"], "r", encoding="utf-8") as handle:
        data = json.load(handle)
    return data if isinstance(data, list) else []


def _optional_form_note(value):
    text = str(value or "").strip()
    if not text:
        return None
    return clean_note(text)


def _scan_students_for_scope(db, user_id, school_year_label, scope):
    sql = "SELECT id, nom_complet, niveau FROM eleves WHERE user_id = ? AND school_year = ?"
    params = [user_id, school_year_label]
    if scope["restricted"] and scope["classes"]:
        placeholders = ",".join("?" * len(scope["classes"]))
        sql += f" AND niveau IN ({placeholders})"
        params += sorted(scope["classes"])
    sql += " ORDER BY niveau, nom_complet"
    return db.execute(sql, params).fetchall()


def _build_scan_preview_rows(extracted_rows, student_rows):
    catalog = build_student_catalog(student_rows)
    preview_rows = []
    for row in extracted_rows:
        match = match_scanned_row(row.get("full_name"), row.get("classe"), catalog)
        issues = row.get("issues") or []
        preview_rows.append(
            {
                "selected": bool(
                    row.get("activite") is not None
                    or row.get("devoir") is not None
                    or row.get("compo") is not None
                    or str(row.get("remarques") or "").strip()
                ),
                "full_name": str(row.get("full_name") or "").strip(),
                "classe": str(row.get("classe") or "").strip(),
                "activite": row.get("activite"),
                "devoir": row.get("devoir"),
                "compo": row.get("compo"),
                "remarques": str(row.get("remarques") or "").strip(),
                "confidence": float(row.get("confidence") or 0),
                "issues": issues,
                "issues_text": "; ".join(str(item) for item in issues if str(item).strip()),
                "match_status": match["status"],
                "match_reason": match["reason"],
                "matched_label": match["matched_label"],
            }
        )
    return preview_rows

@bp.route("/import_excel", methods=["POST"])
@login_required
@write_required
def import_excel():
    user_id = session["user_id"]
    trim = parse_trim(request.form.get("trimestre_import", "1"))
    file = request.files.get("fichier_excel")
    db = get_db()
    selected_school_year = resolve_school_year(
        db,
        request.form.get("school_year"),
        is_admin=bool(session.get("is_admin")),
    )
    scope = (
        {"restricted": False, "subject_ids": set(), "classes": set()}
        if session.get("is_admin")
        else get_user_assignment_scope(db, user_id, selected_school_year)
    )
    if not file or not file.filename:
        flash("Fichier Excel manquant.", "warning")
        return redirect(request.referrer or url_for("dashboard.index", trimestre=trim, school_year=selected_school_year))

    subjects = get_subjects(db, user_id)
    subject_id = select_subject_id(subjects, request.form.get("subject"))
    if scope["restricted"] and subject_id not in scope["subject_ids"]:
        flash("Matiere non autorisee pour ce compte.", "warning")
        return redirect(request.referrer or url_for("dashboard.index", trimestre=trim, school_year=selected_school_year))
    previous_meta = session.get("import_preview")
    if isinstance(previous_meta, dict):
        clear_preview_meta(previous_meta)
    cleanup_import_previews()

    token = uuid.uuid4().hex
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in (".xlsx", ".xls", ".xlsm"):
        ext = ".xlsx"
    preview_path = os.path.join(preview_dir(), f"{token}{ext}")

    try:
        file.save(preview_path)
        all_sheets = pd.read_excel(preview_path, sheet_name=None, header=None)
    except Exception as exc:
        try:
            if os.path.exists(preview_path):
                os.remove(preview_path)
        except Exception:
            pass
        flash(f"Erreur lecture Excel: {exc}", "danger")
        return redirect(request.referrer or url_for("dashboard.index", trimestre=trim, school_year=selected_school_year))

    selected_sheet = ""
    selected_df = None
    header_detected = False
    for sheet_name, raw_df in (all_sheets or {}).items():
        prepared, _, detected = prepare_import_dataframe(raw_df)
        if prepared is None or prepared.empty:
            continue
        selected_sheet = str(sheet_name)
        selected_df = prepared
        header_detected = detected
        break

    if selected_df is None or selected_df.empty:
        try:
            if os.path.exists(preview_path):
                os.remove(preview_path)
        except Exception:
            pass
        flash("Aucune ligne exploitable detectee dans le fichier.", "warning")
        return redirect(request.referrer or url_for("dashboard.index", trimestre=trim, school_year=selected_school_year))

    columns = [str(c) for c in selected_df.columns]
    defaults = build_default_mapping(columns)
    sample_df = selected_df.head(8).copy()
    sample_rows = []
    for _, row in sample_df.iterrows():
        current = {}
        for col in columns:
            value = row.get(col, "")
            if pd.isna(value):
                value = ""
            current[col] = str(value).strip()
        sample_rows.append(current)

    session["import_preview"] = {
        "token": token,
        "path": preview_path,
        "trim": trim,
        "subject_id": int(subject_id),
        "school_year": selected_school_year,
        "sheet_name": selected_sheet,
        "created_at": int(datetime.now().timestamp()),
    }

    if not header_detected:
        flash("Entete non detectee automatiquement: verifiez bien la correspondance des colonnes.", "warning")

    return render_template(
        "import_mapping.html",
        token=token,
        mapping_fields=IMPORT_MAPPING_FIELDS,
        columns=columns,
        defaults=defaults,
        sample_rows=sample_rows,
        sample_headers=columns,
        source_sheet=selected_sheet,
        trim=trim,
        subject_id=subject_id,
        school_year=selected_school_year,
    )


@bp.route("/import_excel_apply", methods=["POST"])
@login_required
@write_required
def import_excel_apply():
    user_id = session["user_id"]
    token = (request.form.get("token") or "").strip()
    meta = get_preview_meta(token)
    if not meta:
        flash("Session d'import expiree. Recommencez l'import.", "warning")
        return redirect(url_for("dashboard.index"))

    trim = parse_trim(meta.get("trim"), "1")
    db = get_db()
    selected_school_year = resolve_school_year(
        db,
        meta.get("school_year"),
        is_admin=bool(session.get("is_admin")),
    )
    scope = (
        {"restricted": False, "subject_ids": set(), "classes": set()}
        if session.get("is_admin")
        else get_user_assignment_scope(db, user_id, selected_school_year)
    )
    subjects = get_subjects(db, user_id)
    subject_ids = {int(s["id"]) for s in subjects}
    try:
        subject_id = int(meta.get("subject_id"))
    except Exception:
        subject_id = None
    if subject_id not in subject_ids:
        subject_id = select_subject_id(subjects, request.form.get("subject"))
    if scope["restricted"] and subject_id not in scope["subject_ids"]:
        clear_preview_meta(meta)
        flash("Matiere non autorisee pour ce compte.", "warning")
        return redirect(url_for("dashboard.index", trimestre=trim, school_year=selected_school_year))

    mapping = {}
    for key, _label in IMPORT_MAPPING_FIELDS:
        mapping[key] = (request.form.get(f"map_{key}") or "").strip()

    try:
        all_sheets = pd.read_excel(meta["path"], sheet_name=None, header=None)
    except Exception as exc:
        clear_preview_meta(meta)
        flash(f"Lecture impossible pendant validation: {exc}", "danger")
        return redirect(url_for("dashboard.index", trimestre=trim, subject=subject_id, school_year=selected_school_year))

    inserted = 0
    updated = 0
    skipped_sheets = 0
    skipped_rows = 0
    use_components = any(
        mapping.get(k)
        for k in ("participation", "comportement", "cahier", "projet", "assiduite_outils")
    )

    for sheet_name, raw_df in (all_sheets or {}).items():
        prepared, _, _ = prepare_import_dataframe(raw_df)
        if prepared is None or prepared.empty:
            skipped_sheets += 1
            continue

        columns = list(prepared.columns)
        resolved = {k: resolve_mapped_column(columns, v) for k, v in mapping.items()}
        if not resolved.get("full_name") and not (
            resolved.get("last_name") or resolved.get("first_name")
        ):
            skipped_sheets += 1
            continue

        for _, row in prepared.iterrows():
            try:
                if resolved.get("full_name"):
                    full = str(row_value(row, resolved["full_name"]) or "").strip()
                else:
                    last_name = str(row_value(row, resolved.get("last_name")) or "").strip()
                    first_name = str(row_value(row, resolved.get("first_name")) or "").strip()
                    full = f"{last_name} {first_name}".strip()

                if not full:
                    skipped_rows += 1
                    continue

                niveau = str(row_value(row, resolved.get("classe")) or "").strip()
                if not niveau:
                    niveau = str(sheet_name).strip() or "Global"
                if scope["restricted"] and niveau not in scope["classes"]:
                    skipped_rows += 1
                    continue

                phone = str(row_value(row, resolved.get("phone")) or "").strip()
                email = str(row_value(row, resolved.get("email")) or "").strip()

                d = clean_note(row_value(row, resolved.get("devoir")))
                c = clean_note(row_value(row, resolved.get("compo")))

                if use_components:
                    p = clean_component(row_value(row, resolved.get("participation")), 3)
                    b = clean_component(row_value(row, resolved.get("comportement")), 6)
                    k = clean_component(row_value(row, resolved.get("cahier")), 5)
                    pr = clean_component(row_value(row, resolved.get("projet")), 4)
                    ao = clean_component(row_value(row, resolved.get("assiduite_outils")), 2)
                else:
                    p, b, k, pr, ao = split_activite_components(
                        row_value(row, resolved.get("activite"))
                    )
                a = sum_activite_components(p, b, k, pr, ao)

                moy = ((d + a) / 2 + (c * 2)) / 3
                rem = get_appreciation_dynamique(moy, user_id)
                custom_rem = row_value(row, resolved.get("remarques"))
                if custom_rem is not None and str(custom_rem).strip():
                    rem = str(custom_rem).strip()

                ex = db.execute(
                    "SELECT id FROM eleves WHERE nom_complet = ? AND niveau = ? AND school_year = ? AND user_id = ?",
                    (full, niveau, selected_school_year, user_id),
                ).fetchone()

                if ex:
                    db.execute(
                        "UPDATE eleves SET parent_phone = COALESCE(?, parent_phone), parent_email = COALESCE(?, parent_email) WHERE id = ?",
                        (phone or None, email or None, ex["id"]),
                    )
                    db.execute(
                        """
                        INSERT INTO notes (
                            user_id, eleve_id, subject_id, trimestre,
                            participation, comportement, cahier, projet, assiduite_outils,
                            activite, devoir, compo, remarques
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(user_id, eleve_id, subject_id, trimestre)
                        DO UPDATE SET
                            participation=excluded.participation,
                            comportement=excluded.comportement,
                            cahier=excluded.cahier,
                            projet=excluded.projet,
                            assiduite_outils=excluded.assiduite_outils,
                            activite=excluded.activite,
                            devoir=excluded.devoir,
                            compo=excluded.compo,
                            remarques=excluded.remarques
                        """,
                        (user_id, ex["id"], subject_id, int(trim), p, b, k, pr, ao, a, d, c, rem),
                    )
                    updated += 1
                else:
                    cur = db.execute(
                        f"INSERT INTO eleves (user_id, school_year, nom_complet, niveau, remarques_t{trim}, devoir_t{trim}, activite_t{trim}, compo_t{trim}, parent_phone, parent_email) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (user_id, selected_school_year, full, niveau, rem, d, a, c, phone, email),
                    )
                    eleve_id = cur.lastrowid
                    db.execute(
                        """
                        INSERT INTO notes (
                            user_id, eleve_id, subject_id, trimestre,
                            participation, comportement, cahier, projet, assiduite_outils,
                            activite, devoir, compo, remarques
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(user_id, eleve_id, subject_id, trimestre)
                        DO UPDATE SET
                            participation=excluded.participation,
                            comportement=excluded.comportement,
                            cahier=excluded.cahier,
                            projet=excluded.projet,
                            assiduite_outils=excluded.assiduite_outils,
                            activite=excluded.activite,
                            devoir=excluded.devoir,
                            compo=excluded.compo,
                            remarques=excluded.remarques
                        """,
                        (user_id, eleve_id, subject_id, int(trim), p, b, k, pr, ao, a, d, c, rem),
                    )
                    inserted += 1
            except Exception:
                skipped_rows += 1
                continue

    db.commit()
    clear_preview_meta(meta)

    total = inserted + updated
    log_change(
        "import_excel",
        user_id,
        details=f"{selected_school_year}: {total} lignes (new {inserted}, upd {updated}, sheets {skipped_sheets}, rows {skipped_rows})",
        subject_id=subject_id,
    )
    flash(
        f"Import termine: {total} lignes (nouveaux {inserted}, maj {updated}, onglets ignores {skipped_sheets}, lignes ignorees {skipped_rows})",
        "success",
    )
    return redirect(url_for("dashboard.index", trimestre=trim, subject=subject_id, school_year=selected_school_year))


@bp.route("/import_scan_pdf", methods=["POST"])
@login_required
@write_required
def import_scan_pdf():
    user_id = session["user_id"]
    trim = parse_trim(request.form.get("trimestre_import_scan", "1"))
    file = request.files.get("fichier_pdf_scan")
    db = get_db()
    selected_school_year = resolve_school_year(
        db,
        request.form.get("school_year"),
        is_admin=bool(session.get("is_admin")),
    )
    scope = (
        {"restricted": False, "subject_ids": set(), "classes": set()}
        if session.get("is_admin")
        else get_user_assignment_scope(db, user_id, selected_school_year)
    )
    if not file or not file.filename:
        flash("PDF scanne manquant.", "warning")
        return redirect(request.referrer or url_for("dashboard.index", trimestre=trim, school_year=selected_school_year))

    subjects = get_subjects(db, user_id)
    subject_id = select_subject_id(subjects, request.form.get("subject"))
    if scope["restricted"] and subject_id not in scope["subject_ids"]:
        flash("Matiere non autorisee pour ce compte.", "warning")
        return redirect(request.referrer or url_for("dashboard.index", trimestre=trim, school_year=selected_school_year))

    previous_meta = session.get(SCAN_PREVIEW_SESSION_KEY)
    if isinstance(previous_meta, dict):
        _clear_scan_preview_meta(previous_meta)
    cleanup_import_previews()

    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext != ".pdf":
        flash("Le scan doit etre fourni au format PDF.", "warning")
        return redirect(request.referrer or url_for("dashboard.index", trimestre=trim, school_year=selected_school_year))

    token = uuid.uuid4().hex
    pdf_path, json_path = _scan_preview_paths(token)
    subject_name = next((str(s["name"]) for s in subjects if int(s["id"]) == int(subject_id)), "")

    try:
        file.save(pdf_path)
        extracted_rows = extract_rows_from_scanned_pdf(
            pdf_path,
            trim=trim,
            subject_name=subject_name,
            school_year=selected_school_year,
        )
        student_rows = _scan_students_for_scope(db, user_id, selected_school_year, scope)
        preview_rows = _build_scan_preview_rows(extracted_rows, student_rows)
        if not preview_rows:
            raise ScanImportError("Aucune ligne exploitable n'a ete detectee dans le PDF scanne.")
        _save_scan_preview_rows(json_path, preview_rows)
        session[SCAN_PREVIEW_SESSION_KEY] = {
            "token": token,
            "path": pdf_path,
            "json_path": json_path,
            "trim": trim,
            "subject_id": int(subject_id),
            "school_year": selected_school_year,
            "filename": os.path.basename(file.filename or "scan.pdf"),
            "created_at": int(datetime.now().timestamp()),
        }
    except ScanImportError as exc:
        _clear_scan_preview_meta({"path": pdf_path, "json_path": json_path})
        flash(str(exc), "warning")
        return redirect(request.referrer or url_for("dashboard.index", trimestre=trim, subject=subject_id, school_year=selected_school_year))
    except Exception as exc:
        _clear_scan_preview_meta({"path": pdf_path, "json_path": json_path})
        flash(f"Erreur pendant l'analyse du scan: {exc}", "danger")
        return redirect(request.referrer or url_for("dashboard.index", trimestre=trim, subject=subject_id, school_year=selected_school_year))

    matched_total = sum(1 for row in preview_rows if row.get("matched_label"))
    return render_template(
        "import_scan_preview.html",
        token=token,
        rows=preview_rows,
        trim=trim,
        subject_id=subject_id,
        subject_name=subject_name,
        school_year=selected_school_year,
        filename=os.path.basename(file.filename or "scan.pdf"),
        detected_total=len(preview_rows),
        matched_total=matched_total,
    )


@bp.route("/import_scan_apply", methods=["POST"])
@login_required
@write_required
def import_scan_apply():
    user_id = session["user_id"]
    token = (request.form.get("token") or "").strip()
    meta = _get_scan_preview_meta(token)
    if not meta:
        flash("Session d'import PDF expiree. Recommencez l'import.", "warning")
        return redirect(url_for("dashboard.index"))

    trim = parse_trim(meta.get("trim"), "1")
    db = get_db()
    selected_school_year = resolve_school_year(
        db,
        meta.get("school_year"),
        is_admin=bool(session.get("is_admin")),
    )
    scope = (
        {"restricted": False, "subject_ids": set(), "classes": set()}
        if session.get("is_admin")
        else get_user_assignment_scope(db, user_id, selected_school_year)
    )
    subjects = get_subjects(db, user_id)
    subject_ids = {int(s["id"]) for s in subjects}
    try:
        subject_id = int(meta.get("subject_id"))
    except Exception:
        subject_id = None
    if subject_id not in subject_ids:
        subject_id = select_subject_id(subjects, request.form.get("subject"))
    if scope["restricted"] and subject_id not in scope["subject_ids"]:
        _clear_scan_preview_meta(meta)
        flash("Matiere non autorisee pour ce compte.", "warning")
        return redirect(url_for("dashboard.index", trimestre=trim, school_year=selected_school_year))

    try:
        base_rows = _load_scan_preview_rows(meta)
    except Exception:
        _clear_scan_preview_meta(meta)
        flash("Impossible de relire la previsualisation du scan.", "danger")
        return redirect(url_for("dashboard.index", trimestre=trim, subject=subject_id, school_year=selected_school_year))

    student_rows = _scan_students_for_scope(db, user_id, selected_school_year, scope)
    catalog = build_student_catalog(student_rows)

    try:
        row_count = int(request.form.get("row_count") or len(base_rows))
    except Exception:
        row_count = len(base_rows)

    updated = 0
    skipped = 0
    unmatched = 0

    for idx in range(row_count):
        if not request.form.get(f"row_{idx}_selected"):
            continue

        full_name = str(request.form.get(f"row_{idx}_full_name") or "").strip()
        classe = str(request.form.get(f"row_{idx}_classe") or "").strip()
        activite = _optional_form_note(request.form.get(f"row_{idx}_activite"))
        devoir = _optional_form_note(request.form.get(f"row_{idx}_devoir"))
        compo = _optional_form_note(request.form.get(f"row_{idx}_compo"))
        remarques = str(request.form.get(f"row_{idx}_remarques") or "").strip()

        if activite is None and devoir is None and compo is None and not remarques:
            skipped += 1
            continue

        match = match_scanned_row(full_name, classe, catalog)
        eleve_id = match.get("student_id")
        if not eleve_id:
            unmatched += 1
            skipped += 1
            continue

        current = db.execute(
            f"""
            SELECT
                e.id,
                e.devoir_t{trim} AS e_devoir,
                e.activite_t{trim} AS e_activite,
                e.compo_t{trim} AS e_compo,
                e.remarques_t{trim} AS e_remarques,
                n.participation,
                n.comportement,
                n.cahier,
                n.projet,
                n.assiduite_outils,
                n.activite AS n_activite,
                n.devoir AS n_devoir,
                n.compo AS n_compo,
                n.remarques AS n_remarques
            FROM eleves e
            LEFT JOIN notes n
              ON n.user_id = e.user_id
             AND n.eleve_id = e.id
             AND n.subject_id = ?
             AND n.trimestre = ?
            WHERE e.id = ? AND e.user_id = ? AND e.school_year = ?
            """,
            (subject_id, int(trim), eleve_id, user_id, selected_school_year),
        ).fetchone()

        if not current:
            skipped += 1
            continue

        current_activite = current["n_activite"] if current["n_activite"] is not None else current["e_activite"]
        current_devoir = current["n_devoir"] if current["n_devoir"] is not None else current["e_devoir"]
        current_compo = current["n_compo"] if current["n_compo"] is not None else current["e_compo"]
        current_remarques = current["n_remarques"] if current["n_remarques"] not in (None, "") else current["e_remarques"]

        final_activite = activite if activite is not None else clean_note(current_activite)
        final_devoir = devoir if devoir is not None else clean_note(current_devoir)
        final_compo = compo if compo is not None else clean_note(current_compo)

        if activite is not None:
            participation, comportement, cahier, projet, assiduite_outils = split_activite_components(final_activite)
        else:
            components = [
                current["participation"],
                current["comportement"],
                current["cahier"],
                current["projet"],
                current["assiduite_outils"],
            ]
            if any(value is not None for value in components):
                participation = clean_component(current["participation"], 3)
                comportement = clean_component(current["comportement"], 6)
                cahier = clean_component(current["cahier"], 5)
                projet = clean_component(current["projet"], 4)
                assiduite_outils = clean_component(current["assiduite_outils"], 2)
            else:
                participation, comportement, cahier, projet, assiduite_outils = split_activite_components(final_activite)

        moyenne = ((final_devoir + final_activite) / 2 + (final_compo * 2)) / 3
        final_remarques = remarques or current_remarques or get_appreciation_dynamique(moyenne, user_id)

        db.execute(
            f"""
            UPDATE eleves
            SET remarques_t{trim} = ?, devoir_t{trim} = ?, activite_t{trim} = ?, compo_t{trim} = ?
            WHERE id = ? AND user_id = ? AND school_year = ?
            """,
            (
                final_remarques,
                final_devoir,
                final_activite,
                final_compo,
                eleve_id,
                user_id,
                selected_school_year,
            ),
        )
        db.execute(
            """
            INSERT INTO notes (
                user_id, eleve_id, subject_id, trimestre,
                participation, comportement, cahier, projet, assiduite_outils,
                activite, devoir, compo, remarques
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(user_id, eleve_id, subject_id, trimestre)
            DO UPDATE SET
                participation=excluded.participation,
                comportement=excluded.comportement,
                cahier=excluded.cahier,
                projet=excluded.projet,
                assiduite_outils=excluded.assiduite_outils,
                activite=excluded.activite,
                devoir=excluded.devoir,
                compo=excluded.compo,
                remarques=excluded.remarques
            """,
            (
                user_id,
                eleve_id,
                subject_id,
                int(trim),
                participation,
                comportement,
                cahier,
                projet,
                assiduite_outils,
                final_activite,
                final_devoir,
                final_compo,
                final_remarques,
            ),
        )
        updated += 1

    db.commit()
    _clear_scan_preview_meta(meta)

    log_change(
        "import_scan_pdf",
        user_id,
        details=f"{selected_school_year}: {updated} lignes maj, {unmatched} non rapprochees, {skipped} ignorees",
        subject_id=subject_id,
    )
    category = "success" if updated else "warning"
    flash(
        f"Import PDF termine: {updated} lignes mises a jour, {unmatched} non rapprochees, {skipped} ignorees.",
        category,
    )
    return redirect(url_for("dashboard.index", trimestre=trim, subject=subject_id, school_year=selected_school_year))

@bp.route("/remplir_bulletin_officiel", methods=["POST"])
@login_required
@write_required
def remplir_bulletin_officiel():
    user_id = session["user_id"]
    trim = parse_trim(request.form.get("trimestre_fill", "1"))
    db = get_db()
    selected_school_year = resolve_school_year(
        db,
        request.form.get("school_year"),
        is_admin=bool(session.get("is_admin")),
    )
    scope = (
        {"restricted": False, "subject_ids": set(), "classes": set()}
        if session.get("is_admin")
        else get_user_assignment_scope(db, user_id, selected_school_year)
    )
    file = request.files.get("fichier_vide")
    if file and file.filename:
        try:
            wb = openpyxl.load_workbook(file)
            subjects = get_subjects(db, user_id)
            subject_id = select_subject_id(subjects, request.form.get("subject"))
            if scope["restricted"] and subject_id not in scope["subject_ids"]:
                flash("Matiere non autorisee pour ce compte.", "warning")
                return redirect(request.referrer or url_for("dashboard.index", trimestre=trim, school_year=selected_school_year))

            for sheet in wb.worksheets:
                header_row = None
                col_map = {}
                for i, row in enumerate(
                    sheet.iter_rows(min_row=1, max_row=20, values_only=True)
                ):
                    row_str = [str(c).lower() for c in row if c]
                    if any(
                        x in row_str
                        for x in ["nom", "\u0627\u0644\u0644\u0642\u0628"]
                    ):
                        header_row = i + 1
                        for cell in sheet[header_row]:
                            if not cell.value:
                                continue
                            v = str(cell.value).strip().lower()
                            if v in ["nom", "\u0627\u0644\u0644\u0642\u0628"]:
                                col_map["nom"] = cell.column
                            elif v in ["prenom", "\u0627\u0644\u0627\u0633\u0645"]:
                                col_map["prenom"] = cell.column
                            elif v in [
                                "01",
                                "1",
                                "act",
                                "\u0627\u0644\u0646\u0634\u0627\u0637",
                                "\u0627\u0644\u0646\u0634\u0627\u0637\u0627\u062a",
                            ]:
                                col_map["act"] = cell.column
                            elif v in [
                                "04",
                                "4",
                                "dev",
                                "\u0627\u0644\u0641\u0631\u0636",
                                "\u0627\u0644\u0648\u0627\u062c\u0628",
                            ]:
                                col_map["dev"] = cell.column
                            elif v in ["09", "9", "compo", "\u0627\u0644\u0627\u062e\u062a\u0628\u0627\u0631"]:
                                col_map["compo"] = cell.column
                            elif v in [
                                "obs",
                                "rem",
                                "remarques",
                                "\u0627\u0644\u062a\u0642\u062f\u064a\u0631\u0627\u062a",
                                "\u0645\u0644\u0627\u062d\u0638\u0627\u062a",
                            ]:
                                col_map["rem"] = cell.column
                        break

                if header_row and "nom" in col_map:
                    for r in range(header_row + 1, sheet.max_row + 1):
                        nom = sheet.cell(row=r, column=col_map["nom"]).value
                        if not nom:
                            continue
                        prenom = (
                            sheet.cell(row=r, column=col_map.get("prenom")).value
                            if col_map.get("prenom")
                            else ""
                        )
                        full = f"{nom} {prenom}".strip()
                        el = db.execute(
                            f"""
                            SELECT
                                e.*,
                                n.activite AS n_activite,
                                n.devoir AS n_devoir,
                                n.compo AS n_compo,
                                n.remarques AS n_remarques
                            FROM eleves e
                            LEFT JOIN notes n
                              ON n.user_id = e.user_id
                             AND n.eleve_id = e.id
                             AND n.subject_id = ?
                             AND n.trimestre = ?
                            WHERE e.nom_complet = ? AND e.user_id = ?
                              AND e.school_year = ?
                            """,
                            (subject_id, int(trim), full, user_id, selected_school_year),
                        ).fetchone()
                        if el:
                            activite_val = (
                                el["n_activite"]
                                if el["n_activite"] is not None
                                else el[f"activite_t{trim}"]
                            )
                            devoir_val = (
                                el["n_devoir"]
                                if el["n_devoir"] is not None
                                else el[f"devoir_t{trim}"]
                            )
                            compo_val = (
                                el["n_compo"]
                                if el["n_compo"] is not None
                                else el[f"compo_t{trim}"]
                            )
                            rem_val = (
                                el["n_remarques"]
                                if el["n_remarques"] not in (None, "")
                                else el[f"remarques_t{trim}"]
                            )
                            if "act" in col_map:
                                sheet.cell(row=r, column=col_map["act"]).value = activite_val
                            if "dev" in col_map:
                                sheet.cell(row=r, column=col_map["dev"]).value = devoir_val
                            if "compo" in col_map:
                                sheet.cell(row=r, column=col_map["compo"]).value = compo_val
                            if "rem" in col_map:
                                sheet.cell(row=r, column=col_map["rem"]).value = rem_val

            out = BytesIO()
            wb.save(out)
            out.seek(0)
            return send_file(
                out,
                download_name="Bulletin_Rempli.xlsx",
                as_attachment=True,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            flash(f"Erreur: {e}", "danger")
    return redirect(request.referrer or url_for("dashboard.index", school_year=selected_school_year))


@bp.route("/import_excel_cancel/<token>")
@login_required
def import_excel_cancel(token: str):
    meta = get_preview_meta((token or "").strip())
    if meta:
        trim = parse_trim(meta.get("trim"), "1")
        school_year = (meta.get("school_year") or "").strip()
        try:
            subject_id = int(meta.get("subject_id"))
        except Exception:
            subject_id = None
        clear_preview_meta(meta)
        if subject_id:
            return redirect(
                url_for(
                    "dashboard.index",
                    trimestre=trim,
                    subject=subject_id,
                    school_year=school_year,
                )
            )
    return redirect(url_for("dashboard.index"))


@bp.route("/import_scan_cancel/<token>")
@login_required
def import_scan_cancel(token: str):
    meta = _get_scan_preview_meta((token or "").strip())
    if meta:
        trim = parse_trim(meta.get("trim"), "1")
        school_year = (meta.get("school_year") or "").strip()
        try:
            subject_id = int(meta.get("subject_id"))
        except Exception:
            subject_id = None
        _clear_scan_preview_meta(meta)
        if subject_id:
            return redirect(
                url_for(
                    "dashboard.index",
                    trimestre=trim,
                    subject=subject_id,
                    school_year=school_year,
                )
            )
    return redirect(url_for("dashboard.index"))

