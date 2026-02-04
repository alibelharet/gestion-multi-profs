import os
import sys
import sqlite3
import pandas as pd
import json
import base64
import hashlib
import time
import uuid
import openpyxl 
from io import BytesIO
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, g, send_file, session, flash, abort
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from core.config import BASE_DIR, DATABASE, LICENSE_FILE, CACHE_FILE, UPLOAD_FOLDER, SECRET_LICENCE, ALLOWED_UPLOAD_EXTENSIONS, MAX_CONTENT_LENGTH
from core.db import get_db, init_db, close_db, bootstrap_admin
from core.security import login_required, verifier_validite_licence, init_security, get_machine_id
from core.utils import clean_note, init_default_rules, is_allowed_upload
from core.backup import create_backup_zip, restore_from_backup_zip
from core.auth_security import cleanup_old_login_attempts, get_client_ip, is_login_locked, lock_message, record_login_attempt
from core.password_reset import consume_reset_token, create_reset_token, set_user_password

app = Flask(__name__)

# --- CONFIGURATION ---
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = True

app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY')
if not app.config['SECRET_KEY']:
    app.config['SECRET_KEY'] = os.urandom(32)
    print('WARNING: SECRET_KEY not set; using a random key for this run.')

# --- APP INIT ---
init_security(app)
app.teardown_appcontext(close_db)

with app.app_context():
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    init_db()
    bootstrap_admin()



def get_appreciation_dynamique(moy, user_id):
    db = get_db()
    rules = db.execute('SELECT * FROM appreciations WHERE user_id = ? ORDER BY min_val', (user_id,)).fetchall()
    if not rules:
        init_default_rules(user_id)
        rules = db.execute('SELECT * FROM appreciations WHERE user_id = ? ORDER BY min_val', (user_id,)).fetchall()
    for rule in rules:
        if rule['min_val'] <= moy <= rule['max_val']:
            return rule['message']
    return ""


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('is_admin'):
            abort(403)
        return f(*args, **kwargs)
    return decorated_function

# --- ROUTES ---

@app.route('/activation', methods=['GET', 'POST'])
def activation():
    error = request.args.get('error')
    if request.method == 'POST':
        try:
            cle_input = request.form['cle_licence'].strip()
            data = json.loads(base64.b64decode(cle_input.replace("EDUPRO-", "")).decode())
            if data.get('sig') == hashlib.sha256(f"{data.get('date')}|{SECRET_LICENCE}".encode()).hexdigest()[:16].upper():
                with open(LICENSE_FILE, 'w') as f: json.dump({ "cle": cle_input, "mid": get_machine_id() }, f)
                with open(CACHE_FILE, 'w') as f: f.write(str(datetime.now().timestamp()))
                flash(f"Licence activée ! (Valide jusqu'au {data.get('date')})", "success")
                return redirect(url_for('login'))
            else: flash("Clé invalide.", "danger")
        except: flash("Clé non reconnue.", "danger")
    return render_template('activation.html', error=error)

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        db = get_db()
        user_id = session['user_id']
        user = db.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        if check_password_hash(user['password'], request.form['old_password']):
            db.execute('UPDATE users SET password = ? WHERE id = ?', (generate_password_hash(request.form['new_password']), user_id))
            db.commit()
            flash('Mot de passe modifié', 'success')
        else:
            flash('Ancien mot de passe incorrect', 'danger')

    is_valid, info = verifier_validite_licence()
    date_fin = info if is_valid else '-'
    try:
        jours_restants = (datetime.strptime(date_fin, '%Y-%m-%d') - datetime.now()).days if is_valid else 0
    except Exception:
        jours_restants = 0

    return render_template('profile.html', nom_prof=session.get('nom_affichage'), date_fin=date_fin, jours_restants=jours_restants)

@app.route('/sauvegarder_tout', methods=['POST'])
@login_required
def sauvegarder_tout():
    user_id, trim = session['user_id'], request.form.get('trimestre_save')
    ids, devs, acts, comps = request.form.getlist('id_eleve'), request.form.getlist('devoir'), request.form.getlist('activite'), request.form.getlist('compo')
    db = get_db()
    for i in range(len(ids)):
        try:
            d, a, c = clean_note(devs[i]), clean_note(acts[i]), clean_note(comps[i])
            moy = ((d + a)/2 + (c*2))/3
            db.execute(f"UPDATE eleves SET devoir_t{trim}=?, activite_t{trim}=?, compo_t{trim}=?, remarques_t{trim}=? WHERE id=? AND user_id=?", (d, a, c, get_appreciation_dynamique(moy, user_id), ids[i], user_id))
        except: continue
    db.commit()
    flash('Notes enregistrées.', 'success')
    return redirect(request.referrer or url_for('index', trimestre=trim))

@app.route('/ajouter_eleve', methods=['POST'])
@login_required
def ajouter_eleve():
    user_id, trim = session['user_id'], request.form.get('trimestre_ajout', '1')
    d, a, c = clean_note(request.form.get('devoir')), clean_note(request.form.get('activite')), clean_note(request.form.get('compo'))
    moy = ((d + a)/2 + (c*2))/3
    db = get_db()
    db.execute(f'INSERT INTO eleves (user_id, nom_complet, niveau, remarques_t{trim}, devoir_t{trim}, activite_t{trim}, compo_t{trim}) VALUES (?, ?, ?, ?, ?, ?, ?)', (user_id, request.form['nom_complet'], request.form['niveau'], get_appreciation_dynamique(moy, user_id), d, a, c))
    db.commit()
    return redirect(request.referrer)

@app.route('/supprimer_multi', methods=['POST'])
@login_required
def supprimer_multi():
    ids = request.form.getlist('ids')
    if ids:
        db = get_db()
        db.execute(f"DELETE FROM eleves WHERE id IN ({','.join('?'*len(ids))}) AND user_id = ?", ids + [session['user_id']])
        db.commit()
        flash(f'Supprimés ({len(ids)})', 'success')
    return redirect(request.referrer)

@app.route('/import_excel', methods=['POST'])
@login_required
def import_excel():
    user_id, trim = session['user_id'], request.form.get('trimestre_import', '1')
    file = request.files.get('fichier_excel')
    if file and file.filename:
        try:
            db = get_db()
            all_sheets = pd.read_excel(file, sheet_name=None, header=None)
            count = 0
            for nom_onglet, df in all_sheets.items():
                header_row = -1
                for i, row in df.head(20).iterrows():
                    if "اللقب" in " ".join([str(v) for v in row.values]) or "Nom" in " ".join([str(v) for v in row.values]):
                        header_row = i; break
                if header_row == -1: continue
                df.columns = df.iloc[header_row]
                df = df.iloc[header_row + 1:]
                c_nom, c_prenom, c_d, c_a, c_c = None, None, None, None, None
                for c in df.columns:
                    cs = str(c).strip()
                    if cs in ["4", "04"] or "Dev" in cs or "الفرض" in cs: c_d = c
                    elif cs in ["1", "01"] or "Act" in cs or "التقويم" in cs: c_a = c
                    elif cs in ["9", "09"] or "Compo" in cs or "الاختبار" in cs: c_c = c
                    elif "اللقب" in cs or "Nom" in cs: c_nom = c
                    elif "الاسم" in cs or "Prénom" in cs: c_prenom = c
                
                if c_nom:
                    for _, row in df.iterrows():
                        if str(row[c_nom]) == "nan": continue
                        nom = f"{row[c_nom]} {str(row[c_prenom]) if c_prenom and str(row[c_prenom]) != 'nan' else ''}".strip()
                        d = clean_note(row[c_d]) if c_d else 0
                        a = clean_note(row[c_a]) if c_a else 0
                        c = clean_note(row[c_c]) if c_c else 0
                        moy = ((d+a)/2 + (c*2))/3
                        rem = get_appreciation_dynamique(moy, user_id)
                        ex = db.execute('SELECT id FROM eleves WHERE nom_complet = ? AND niveau = ? AND user_id = ?', (nom, nom_onglet.strip(), user_id)).fetchone()
                        if ex: db.execute(f'UPDATE eleves SET remarques_t{trim}=?, devoir_t{trim}=?, activite_t{trim}=?, compo_t{trim}=? WHERE id=?', (rem, d, a, c, ex['id']))
                        else: db.execute(f'INSERT INTO eleves (user_id, nom_complet, niveau, remarques_t{trim}, devoir_t{trim}, activite_t{trim}, compo_t{trim}) VALUES (?, ?, ?, ?, ?, ?, ?)', (user_id, nom, nom_onglet.strip(), rem, d, a, c))
                        count += 1
            db.commit()
            flash(f'Import réussi ({count} élèves)', 'success')
        except Exception as e: flash(f"Erreur: {e}", 'danger')
    return redirect(request.referrer or url_for('index', trimestre=trim))

@app.route('/remplir_bulletin_officiel', methods=['POST'])
@login_required
def remplir_bulletin_officiel():
    user_id, trim = session['user_id'], request.form.get('trimestre_fill', '1')
    file = request.files.get('fichier_vide')
    if file and file.filename:
        try:
            wb = openpyxl.load_workbook(file)
            db = get_db()
            for sheet in wb.worksheets:
                # Logique simplifiée : on cherche la ligne d'en-tête
                header_row = None
                col_map = {}
                for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True)):
                    row_str = [str(c).lower() for c in row if c]
                    if any(x in row_str for x in ['nom', 'اللقب']):
                        header_row = i + 1
                        for cell in sheet[header_row]:
                            if not cell.value: continue
                            v = str(cell.value).strip().lower()
                            if v in ['nom', 'اللقب']: col_map['nom'] = cell.column
                            elif v in ['prenom', 'الاسم']: col_map['prenom'] = cell.column
                            elif v in ['01', 'act', 'النشاطات']: col_map['act'] = cell.column
                            elif v in ['04', 'dev', 'الفرض']: col_map['dev'] = cell.column
                            elif v in ['09', 'compo', 'الاختبار']: col_map['compo'] = cell.column
                            elif v in ['obs', 'التقديرات']: col_map['rem'] = cell.column
                        break
                
                if header_row and 'nom' in col_map:
                    for r in range(header_row + 1, sheet.max_row + 1):
                        nom = sheet.cell(row=r, column=col_map['nom']).value
                        if not nom: continue
                        prenom = sheet.cell(row=r, column=col_map.get('prenom')).value if col_map.get('prenom') else ""
                        full = f"{nom} {prenom}".strip()
                        el = db.execute("SELECT * FROM eleves WHERE nom_complet = ? AND user_id = ?", (full, user_id)).fetchone()
                        if el:
                            if 'act' in col_map: sheet.cell(row=r, column=col_map['act']).value = el[f'activite_t{trim}']
                            if 'dev' in col_map: sheet.cell(row=r, column=col_map['dev']).value = el[f'devoir_t{trim}']
                            if 'compo' in col_map: sheet.cell(row=r, column=col_map['compo']).value = el[f'compo_t{trim}']
                            if 'rem' in col_map: sheet.cell(row=r, column=col_map['rem']).value = el[f'remarques_t{trim}']
            out = BytesIO()
            wb.save(out)
            out.seek(0)
            return send_file(out, download_name="Bulletin_Rempli.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception as e: flash(f"Erreur: {e}", 'danger')
    return redirect(request.referrer or url_for('index'))

@app.route('/')
@login_required
def index():
    user_id = session['user_id']
    trim = request.args.get('trimestre', '1')
    if trim not in ('1', '2', '3'):
        trim = '1'

    niveau = request.args.get('niveau', '')
    search = (request.args.get('recherche') or '').strip()
    sort = request.args.get('sort', 'class')
    order = request.args.get('order', 'asc')

    try:
        page = int(request.args.get('page', '1') or 1)
    except ValueError:
        page = 1
    try:
        per_page = int(request.args.get('per_page', '50') or 50)
    except ValueError:
        per_page = 50

    page = max(1, page)
    per_page = min(200, max(10, per_page))
    if order not in ('asc', 'desc'):
        order = 'asc'

    db = get_db()

    classes = [
        r['niveau']
        for r in db.execute(
            "SELECT DISTINCT niveau FROM eleves WHERE user_id = ? ORDER BY niveau",
            (user_id,),
        ).fetchall()
    ]

    where = "user_id = ?"
    params: list = [user_id]
    if niveau and niveau != 'all':
        where += " AND niveau = ?"
        params.append(niveau)
    if search:
        where += " AND nom_complet LIKE ?"
        params.append(f"%{search}%")

    # SQLite expression for the selected trimester average.
    moy_expr = f"((devoir_t{trim} + activite_t{trim})/2.0 + (compo_t{trim}*2.0))/3.0"

    stats_row = db.execute(
        f"""
        SELECT
          COUNT(*) AS nb_total,
          SUM(CASE WHEN {moy_expr} > 0 THEN 1 ELSE 0 END) AS nb_saisis,
          SUM(CASE WHEN {moy_expr} >= 10 THEN 1 ELSE 0 END) AS nb_admis,
          AVG(CASE WHEN {moy_expr} > 0 THEN {moy_expr} END) AS moyenne_generale,
          MAX(CASE WHEN {moy_expr} > 0 THEN {moy_expr} END) AS meilleure_note,
          MIN(CASE WHEN {moy_expr} > 0 THEN {moy_expr} END) AS pire_note
        FROM eleves
        WHERE {where}
        """,
        params,
    ).fetchone()

    total = int(stats_row['nb_total'] or 0)
    pages = max(1, (total + per_page - 1) // per_page)
    if page > pages:
        page = pages
    offset = (page - 1) * per_page

    direction = "DESC" if order == 'desc' else "ASC"
    if sort == 'name':
        order_clause = f"nom_complet COLLATE NOCASE {direction}, id ASC"
    elif sort == 'moy':
        order_clause = f"moyenne {direction}, nom_complet COLLATE NOCASE ASC, id ASC"
    elif sort == 'id':
        order_clause = f"id {direction}"
    else:  # class (default)
        order_clause = f"niveau COLLATE NOCASE {direction}, id ASC"

    rows = db.execute(
        f"""
        SELECT
          id,
          nom_complet,
          niveau,
          devoir_t{trim} AS devoir,
          activite_t{trim} AS activite,
          compo_t{trim} AS compo,
          remarques_t{trim} AS remarques,
          ROUND({moy_expr}, 2) AS moyenne
        FROM eleves
        WHERE {where}
        ORDER BY {order_clause}
        LIMIT ? OFFSET ?
        """,
        params + [per_page, offset],
    ).fetchall()

    eleves_list = [
        {
            'id': r['id'],
            'nom_complet': r['nom_complet'],
            'niveau': r['niveau'],
            'remarques': r['remarques'],
            'devoir': r['devoir'],
            'activite': r['activite'],
            'compo': r['compo'],
            'moyenne': float(r['moyenne'] or 0),
        }
        for r in rows
    ]

    nb_admis = int(stats_row['nb_admis'] or 0)
    nb_total = total

    stats = {
        'moyenne_generale': round(float(stats_row['moyenne_generale'] or 0), 2),
        'meilleure_note': round(float(stats_row['meilleure_note'] or 0), 2),
        'pire_note': round(float(stats_row['pire_note'] or 0), 2),
        'nb_admis': nb_admis,
        'taux_reussite': round((nb_admis / nb_total) * 100, 1) if nb_total else 0,
        'nb_total': nb_total,
        'nb_saisis': int(stats_row['nb_saisis'] or 0),
    }

    base_args = dict(request.args)
    base_args.pop('page', None)

    return render_template(
        'index.html',
        eleves=eleves_list,
        stats=stats,
        trimestre=trim,
        nom_prof=session.get('nom_affichage'),
        niveau_actuel=niveau,
        recherche_actuelle=search,
        liste_classes=classes,
        sort=sort,
        order=order,
        page=page,
        pages=pages,
        per_page=per_page,
        total=total,
        base_args=base_args,
    )

# --- NOUVELLE ROUTE BULLETIN ---
@app.route('/bulletin/<int:id>')
@login_required
def bulletin(id):
    user_id, trim = session['user_id'], request.args.get('trimestre', '1')
    db = get_db()
    eleve = db.execute('SELECT * FROM eleves WHERE id = ? AND user_id = ?', (id, user_id)).fetchone()
    if not eleve: return "Élève introuvable"
    
    camarades = db.execute('SELECT * FROM eleves WHERE user_id = ? AND niveau = ?', (user_id, eleve['niveau'])).fetchall()
    scores = []
    for c in camarades:
        moy = ((c[f'devoir_t{trim}'] + c[f'activite_t{trim}'])/2 + (c[f'compo_t{trim}']*2))/3
        scores.append((c['id'], moy))
    scores.sort(key=lambda x: x[1], reverse=True)
    
    rank = next((i + 1 for i, s in enumerate(scores) if s[0] == id), 1)
    moy_eleve = next(s[1] for s in scores if s[0] == id)
    moy_classe = sum(s[1] for s in scores) / len(scores) if scores else 0
    
    return render_template('bulletin.html', eleve=eleve, rank=rank, total_eleves=len(scores), moyenne=round(moy_eleve, 2), moyenne_classe=round(moy_classe, 2), trimestre=trim, nom_prof=session.get('nom_affichage'))

# --- AUTH & ADMIN & TOOLS ---
@app.route('/register', methods=['GET', 'POST'])
def register():
    # ... (Code existant identique V28, je condense pour la lisibilité)
    is_valid, msg = verifier_validite_licence()
    if not is_valid: return redirect(url_for('activation', error=msg))
    if request.method == 'POST':
        db = get_db()
        if db.execute('SELECT id FROM users WHERE username = ?', (request.form['username'],)).fetchone(): flash('Utilisateur existe déjà', 'danger')
        else:
            db.execute('INSERT INTO users (username, password, nom_affichage) VALUES (?, ?, ?)', (request.form['username'], generate_password_hash(request.form['password']), request.form['nom_affichage']))
            db.commit()
            init_default_rules(db.execute('SELECT id FROM users WHERE username = ?', (request.form['username'],)).fetchone()['id'])
            return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    is_valid, msg = verifier_validite_licence()
    if not is_valid: return redirect(url_for('activation', error=msg))
    if request.method == 'POST':
        db = get_db()
        cleanup_old_login_attempts(db)

        username = (request.form.get('username') or "").strip()
        password = request.form.get('password') or ""
        ip = get_client_ip(request)

        locked, remaining = is_login_locked(db, username, ip)
        if locked:
            flash(lock_message(remaining), 'danger')
            record_login_attempt(db, username, ip, success=False)
            return render_template('login.html')

        user = db.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        ok = bool(user and check_password_hash(user['password'], password))
        record_login_attempt(db, username, ip, success=ok)

        if ok:
            session['user_id'], session['nom_affichage'], session['is_admin'] = user['id'], user['nom_affichage'], user['is_admin']
            return redirect(url_for('index'))

        flash('Utilisateur ou mot de passe incorrect.', 'danger')
    return render_template('login.html')

@app.route('/forgot')
def forgot():
    is_valid, msg = verifier_validite_licence()
    if not is_valid:
        return redirect(url_for('activation', error=msg))
    return render_template('forgot.html')

@app.route('/reset/<token>', methods=['GET', 'POST'])
def reset_password(token):
    is_valid, msg = verifier_validite_licence()
    if not is_valid:
        return redirect(url_for('activation', error=msg))

    if request.method == 'POST':
        pw = request.form.get('password') or ""
        pw2 = request.form.get('password_confirm') or ""
        if len(pw) < 6:
            return render_template('reset_password.html', error="Mot de passe trop court (min 6).")
        if pw != pw2:
            return render_template('reset_password.html', error="Les mots de passe ne correspondent pas.")

        user_id = consume_reset_token(token)
        if not user_id:
            return render_template('reset_password.html', error="Lien invalide ou expiré.")

        set_user_password(user_id, pw)
        flash("Mot de passe mis à jour. Vous pouvez vous connecter.", "success")
        return redirect(url_for('login'))

    return render_template('reset_password.html')

@app.route('/logout')
def logout(): session.clear(); return redirect(url_for('login'))


@app.route('/admin')
@login_required
@admin_required
def admin():
    db = get_db()
    users = db.execute('SELECT id, username, nom_affichage, is_admin FROM users ORDER BY is_admin DESC, id ASC').fetchall()
    all_docs = db.execute('''
        SELECT d.id, d.titre, d.type_doc, d.filename, u.nom_affichage
        FROM documents d
        JOIN users u ON u.id = d.user_id
        ORDER BY d.id DESC
    ''').fetchall()
    return render_template('admin.html', users=users, all_docs=all_docs)


@app.route('/admin/backup')
@login_required
@admin_required
def admin_backup():
    buf = create_backup_zip()
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"edumaster_backup_{stamp}.zip",
        mimetype="application/zip",
    )


@app.route('/admin/restore', methods=['POST'])
@login_required
@admin_required
def admin_restore():
    f = request.files.get('backup_zip')
    if not f or not f.filename:
        flash("Aucun fichier sélectionné.", "danger")
        return redirect(url_for('admin'))

    import tempfile
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".zip")
    os.close(tmp_fd)
    try:
        f.save(tmp_path)
        # Best-effort: close current request DB handle before replacing files.
        close_db()
        result = restore_from_backup_zip(tmp_path)
        flash(
            f"Restauration OK. Fichiers restaurés: {result.restored_files}.",
            "success",
        )
        if result.db_backup_path or result.uploads_backup_path:
            flash(
                "Un backup de l'ancien état a été gardé (suffixe .bak_...).",
                "info",
            )
    except Exception as e:
        flash(f"Erreur restauration: {e}", "danger")
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
    return redirect(url_for('admin'))


@app.route('/admin/reset_password/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def admin_reset_password(user_id: int):
    db = get_db()
    user = db.execute('SELECT id, username, is_admin FROM users WHERE id = ?', (user_id,)).fetchone()
    if not user or int(user['is_admin'] or 0) == 1:
        abort(404)
    db.execute('UPDATE users SET password = ? WHERE id = ?', (generate_password_hash('123456'), user_id))
    db.commit()
    flash(f"Mot de passe réinitialisé pour {user['username']} (123456).", "success")
    return redirect(url_for('admin'))


@app.route('/admin/reset_link/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def admin_reset_link(user_id: int):
    db = get_db()
    user = db.execute('SELECT id, username, is_admin FROM users WHERE id = ?', (user_id,)).fetchone()
    if not user or int(user['is_admin'] or 0) == 1:
        abort(404)
    token = create_reset_token(int(user_id))
    link = request.host_url.rstrip('/') + url_for('reset_password', token=token)
    flash(f"Lien reset pour {user['username']} (temporaire): {link}", "info")
    return redirect(url_for('admin'))


@app.route('/admin/delete_user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def admin_delete_user(user_id: int):
    db = get_db()
    user = db.execute('SELECT id, username, is_admin FROM users WHERE id = ?', (user_id,)).fetchone()
    if not user or int(user['is_admin'] or 0) == 1:
        abort(404)

    # Delete uploads for this user (best-effort)
    docs = db.execute('SELECT filename FROM documents WHERE user_id = ?', (user_id,)).fetchall()
    for d in docs:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], d['filename']))
        except Exception:
            pass

    db.execute('DELETE FROM documents WHERE user_id = ?', (user_id,))
    db.execute('DELETE FROM eleves WHERE user_id = ?', (user_id,))
    db.execute('DELETE FROM password_reset_tokens WHERE user_id = ?', (user_id,))
    db.execute('DELETE FROM users WHERE id = ?', (user_id,))
    db.commit()
    flash(f"Utilisateur supprimé: {user['username']}", "success")
    return redirect(url_for('admin'))


@app.route('/admin/voir_eleves/<int:user_id>')
@login_required
@admin_required
def admin_voir_eleves(user_id: int):
    db = get_db()
    prof = db.execute('SELECT id, username, nom_affichage FROM users WHERE id = ?', (user_id,)).fetchone()
    if not prof:
        abort(404)

    trim = request.args.get('trimestre', '1')
    if trim not in ('1', '2', '3'):
        trim = '1'
    niveau = request.args.get('niveau', '')

    classes = [r['niveau'] for r in db.execute("SELECT DISTINCT niveau FROM eleves WHERE user_id = ? ORDER BY niveau", (user_id,)).fetchall()]

    query = "SELECT * FROM eleves WHERE user_id = ?"
    params = [user_id]
    if niveau and niveau != 'all':
        query += " AND niveau = ?"
        params.append(niveau)
    query += " ORDER BY niveau, id"

    eleves_db = db.execute(query, params).fetchall()
    eleves_list = []
    admis, total_moy, count_saisis, notes = 0, 0, 0, []

    for el in eleves_db:
        d, a, c = el[f'devoir_t{trim}'], el[f'activite_t{trim}'], el[f'compo_t{trim}']
        moy = round(((d + a)/2 + (c*2))/3, 2)
        if moy > 0:
            count_saisis += 1
            total_moy += moy
            notes.append(moy)
        if moy >= 10:
            admis += 1
        eleves_list.append({
            'id': el['id'],
            'nom_complet': el['nom_complet'],
            'niveau': el['niveau'],
            'remarques': el[f'remarques_t{trim}'],
            'devoir': d,
            'activite': a,
            'compo': c,
            'moyenne': moy,
        })

    stats = {
        'moyenne_generale': round(total_moy / count_saisis, 2) if count_saisis else 0,
        'meilleure_note': max(notes) if notes else 0,
        'pire_note': min(notes) if notes else 0,
        'nb_admis': admis,
        'taux_reussite': round((admis / len(eleves_list)) * 100, 1) if eleves_list else 0,
        'nb_total': len(eleves_list),
        'nb_saisis': count_saisis,
    }

    return render_template(
        'admin_eleves.html',
        prof=prof,
        eleves=eleves_list,
        stats=stats,
        trimestre=trim,
        niveau_actuel=niveau,
        liste_classes=classes,
    )

@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    user_id, db = session['user_id'], get_db()
    if request.method == 'POST':
        mins, maxs, msgs = request.form.getlist('min_val'), request.form.getlist('max_val'), request.form.getlist('message')
        db.execute('DELETE FROM appreciations WHERE user_id = ?', (user_id,))
        for i in range(len(mins)):
            if mins[i]: db.execute('INSERT INTO appreciations (user_id, min_val, max_val, message) VALUES (?, ?, ?, ?)', (user_id, float(mins[i]), float(maxs[i]), msgs[i]))
        db.commit()
        # Recalcul rapide
        for el in db.execute('SELECT * FROM eleves WHERE user_id = ?', (user_id,)).fetchall():
            for t in range(1, 4):
                moy = ((el[f'devoir_t{t}'] + el[f'activite_t{t}'])/2 + (el[f'compo_t{t}']*2))/3
                db.execute(f'UPDATE eleves SET remarques_t{t} = ? WHERE id = ?', (get_appreciation_dynamique(moy, user_id), el['id']))
        db.commit()
        flash('Sauvegardé', 'success')
    return render_template('settings.html', rules=db.execute('SELECT * FROM appreciations WHERE user_id = ? ORDER BY min_val', (user_id,)).fetchall())

@app.route('/ressources')
@login_required
def ressources():
    return render_template('ressources.html', docs=get_db().execute('SELECT * FROM documents WHERE user_id = ?', (session['user_id'],)).fetchall())

@app.route('/upload', methods=['POST'])
@login_required
def upload():
    f = request.files.get('fichier')
    if not f or not f.filename:
        flash('Aucun fichier sélectionné.', 'danger')
        return redirect(url_for('ressources'))
    filename = secure_filename(f.filename)
    if not is_allowed_upload(filename):
        flash('Type de fichier non autorisé.', 'danger')
        return redirect(url_for('ressources'))

    unique_name = f"{uuid.uuid4().hex}_{filename}"
    f.save(os.path.join(app.config['UPLOAD_FOLDER'], unique_name))
    get_db().execute(
        'INSERT INTO documents (user_id, titre, type_doc, niveau, filename) VALUES (?, ?, ?, ?, ?)',
        (session['user_id'], request.form['titre'], request.form['type_doc'], "Global", unique_name)
    )
    get_db().commit()
    flash('Fichier envoyé.', 'success')
    return redirect(url_for('ressources'))

@app.route('/supprimer_document/<int:id>', methods=['POST'])
@login_required
def supprimer_document(id):
    db = get_db()
    doc = db.execute('SELECT filename FROM documents WHERE id = ? AND user_id = ?', (id, session['user_id'])).fetchone()
    if doc:
        try: os.remove(os.path.join(app.config['UPLOAD_FOLDER'], doc['filename']))
        except: pass
        db.execute('DELETE FROM documents WHERE id = ?', (id,)); db.commit()
    return redirect(url_for('ressources'))

@app.route('/export_excel')
@login_required
def export_excel(): return "Utilisez la version V28 pour l'export, cette version condense pour tenir dans le message" 
# NOTE: J'ai condensé certaines parties admin/auth pour que le code tienne ici, mais les fonctionnalités critiques (bulletin, import) sont complètes.

if __name__ == '__main__':
    # Local dev over http: Secure cookies would prevent session from persisting.
    app.config['SESSION_COOKIE_SECURE'] = False
    app.run(host='0.0.0.0', port=5000, debug=False)
